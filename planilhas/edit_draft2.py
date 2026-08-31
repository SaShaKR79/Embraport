#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inclui no draft2 da MLA (aba "base 30 06 26com dividendos") os dois blocos das
páginas impressas de 31.08: "1. P. Líquido em 31.07.26" (PL 254.032 em R$ mil,
EV 5.954.811 e quadro 60/40 × 34%) e "2. Apuração da Carga IR" (custo 66.556 =
26,2% × PL; cash out 1.200.000; pgto agora 385.371). Tudo por fórmula, no estilo
do próprio draft (labels em negrito, sublinhados, amarelo = input). Nada do
conteúdo existente é alterado — os blocos são acrescentados abaixo (a partir da
linha 139), sem deslocar linhas.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "uploads/draft2_mla.xlsx"
OUT = "Exercicio_Incorporacao_draft2_MLA_31.08_v2.xlsx"

wb = openpyxl.load_workbook(SRC)
ws = wb["base 30 06 26com dividendos"]

f_b = Font(name="Calibri", size=11, bold=True)
f_t = Font(name="Calibri", size=11)
f_i = Font(name="Calibri", size=11, italic=True)
f_bi = Font(name="Calibri", size=11, bold=True, italic=True)
yel = PatternFill("solid", fgColor="FFFFFF00")
thin = Side(style="thin", color="FF000000")
med = Side(style="medium", color="FF000000")
b_bottom = Border(bottom=thin)
b_top = Border(top=thin)
b_double = Border(top=thin, bottom=Side(style="double", color="FF000000"))
al_r = Alignment(horizontal="right")
al_c = Alignment(horizontal="center")
NFI = "#,##0;(#,##0)"
PCT = "0.00%"

def put(coord, value, font=f_t, fill=None, nf=None, al=None, border=None):
    c = ws[coord]
    c.value = value
    c.font = font
    if fill: c.fill = fill
    if nf: c.number_format = nf
    if al: c.alignment = al
    if border: c.border = border
    return c

# ============ 1. P. Líquido em 31.07.26 ============
put("D139", "1. P. Liquido em 31.07.26  (R$ mil)", font=f_b)
for col in "DEFGHIJK":
    ws[f"{col}139"].border = b_bottom
    if col != "D": ws[f"{col}139"].font = f_b

put("D141", "Embracon", font=f_t)
put("G141", 254032.0, fill=yel, nf=NFI, al=al_r)

put("D143", "Savian", font=f_t); put("E143", "investimento", font=f_i)
put("G143", "=G141/2", nf=NFI, al=al_r)
put("H143", "=G143/G$145", nf=PCT, al=al_r)
put("D144", "JVFJ", font=f_t); put("E144", "investimento", font=f_i)
put("G144", "=G141/2", nf=NFI, al=al_r, border=b_bottom)
put("H144", "=G144/G$145", nf=PCT, al=al_r)
put("G145", "=G143+G144", nf=NFI, al=al_r)

put("D148", "Enterprise Value", font=f_b)
put("D150", "Embracon", font=f_t)
put("G150", 5132887.0, fill=yel, nf=NFI, al=al_r)
put("H150", "=G150/G$152", nf=PCT, al=al_r)
put("D151", "CNP", font=f_t)
put("G151", 821924.0, fill=yel, nf=NFI, al=al_r, border=b_bottom)
put("H151", "=G151/G$152", nf=PCT, al=al_r, border=b_bottom)
put("G152", "=G150+G151", nf=NFI, al=al_r)
put("H152", "=G152/G152", nf=PCT, al=al_r)
put("I152", '"ppA"', font=f_t)

# quadro 60/40
put("F154", "=G152", nf=NFI, al=al_c)
put("F156", "EMBRACON", font=f_b, al=al_c)
put("H156", "CNP", font=f_b, al=al_c)
put("F157", 0.60, fill=yel, nf="0%", al=al_c)
put("H157", "=1-F157", nf="0%", al=al_c)
put("F158", "=$G$152*F157", nf=NFI, al=al_c)
put("H158", "=$G$152*H157", nf=NFI, al=al_c)
put("F159", "=$I$96", nf="0%", al=al_c)
put("H159", "=$I$96", nf="0%", al=al_c)
put("F160", "=F158*F159", font=f_bi, nf=NFI, al=al_c)
put("H160", "=H158*H159", font=f_bi, nf=NFI, al=al_c)
# moldura do quadro
for r in range(154, 161):
    for col in "FGH":
        cell = ws[f"{col}{r}"]
        left = med if col == "F" else None
        right = med if col == "H" else None
        top = med if r == 154 else (thin if r == 158 else None)
        bottom = med if r == 160 else (thin if r == 157 else None)
        cell.border = Border(left=left, right=right, top=top, bottom=bottom)

# ============ 2. Apuração da Carga IR ============
put("D163", "2. Apuração da Carga IR  (R$ mil)", font=f_b)
for col in "DEFGHIJK":
    ws[f"{col}163"].border = b_bottom
    if col != "D": ws[f"{col}163"].font = f_b

put("J164", "P.Liq E", font=f_t, al=al_r)
put("D165", "Custo Investimento", font=f_t)
put("G165", "=-J165*$E$88", nf=NFI, al=al_r)
put("H165", "<--------------", font=f_t)
put("J165", "=G141", nf=NFI, al=al_r)
put("D166", "Cash Out", font=f_t)
put("G166", 1200000.0, fill=yel, nf=NFI, al=al_r, border=b_bottom)
put("D167", "ganho", font=f_t)
put("G167", "=G166+G165", nf=NFI, al=al_r)
put("D168", "IR", font=f_t)
put("G168", "=$I$96", nf="0.00%", al=al_r, border=b_bottom)
put("D169", "pgto agora", font=f_b)
put("G169", "=G167*G168", font=f_b, nf=NFI, al=al_r, border=b_double)
put("H169", "", border=b_double)

put("D171", "Custo alocado = 26,2% (E88) × P.Líq Embracon 31.07.26 — critério do draft (custo × p.p. vendidos). Blocos em R$ mil; demais seções da aba em R$ milhões (base 30.06.26).", font=f_i)

wb.save(OUT)
print("saved", OUT)
EOF_MARKER_NOT_USED = None
