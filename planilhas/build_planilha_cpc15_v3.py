#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
v3 — inclui a seção "Situação após a incorporação — acervo vertido (cindido)"
na aba Contabilização (nome dado pelo usuário na v3), entre a incorporação e o
cash out, espelhando o bloco "SITUAÇÃO APÓS INCORPORAÇÃO" do draft MLA.
Demais abas/blocos idênticos à v3 do usuário.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image

OUT = "Exercicio_Incorporacao_CPC15_v3.xlsx"
LOGO = "xlsx_media_modelo/image1.png"

RED = "FFC00000"; ZEBRA = "FFEAF0FB"; YEL = "FFFFF2CC"; GRAY_B = "FFC9C9C9"
f_title = Font(name="Calibri", size=12, bold=True)
f_bar = Font(name="Calibri", size=9, bold=True, color="FFFFFFFF")
f_b = Font(name="Calibri", size=9, bold=True)
f_t = Font(name="Calibri", size=9)
f_i = Font(name="Calibri", size=9, italic=True)
fill_red = PatternFill("solid", fgColor=RED)
fill_zebra = PatternFill("solid", fgColor=ZEBRA)
fill_yel = PatternFill("solid", fgColor=YEL)
thin = Side(style="thin", color=GRAY_B)
b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
al_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
al_l = Alignment(horizontal="left", vertical="center", wrap_text=True)
al_r = Alignment(horizontal="right", vertical="center")
NF = "#,##0.0;[Red]\\(#,##0.0\\)"
PCT = "0.0%"; PCT2 = "0.00%"

wb = openpyxl.Workbook()

def put(ws, coord, value, font=f_t, fill=None, nf=None, al=None, border=b_all):
    c = ws[coord]
    c.value = value; c.font = font
    if fill: c.fill = fill
    if nf: c.number_format = nf
    c.alignment = al or (al_r if nf else al_l)
    if border: c.border = border
    return c

def bar(ws, rng, text):
    ws.merge_cells(rng)
    r1, r2 = rng.split(":")
    row = int("".join(ch for ch in r1 if ch.isdigit()))
    c1 = openpyxl.utils.column_index_from_string("".join(ch for ch in r1 if ch.isalpha()))
    c2 = openpyxl.utils.column_index_from_string("".join(ch for ch in r2 if ch.isalpha()))
    put(ws, r1, text, font=f_bar, fill=fill_red, al=al_c)
    for cc in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=cc); cell.border = b_all; cell.fill = fill_red

def hdr(ws, row, cols_labels, first_label=""):
    put(ws, f"B{row}", first_label, font=f_b, fill=fill_zebra)
    for col, lbl in cols_labels:
        put(ws, f"{col}{row}", lbl, font=f_b, fill=fill_zebra, al=al_c)

def blank(ws, row, cols):
    for col in cols:
        put(ws, f"{col}{row}", "", border=b_all)

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

# =====================================================================
# ABA 1 — Premissas (idêntica à v3)
# =====================================================================
ws = wb.active
ws.title = "Premissas"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 50, "C": 13, "D": 52})
try:
    img = Image(LOGO); img.width, img.height = 147, 35
    ws.add_image(img, "B1")
except Exception:
    pass
put(ws, "E1", "Projeto Another House — Exercício da Incorporação sob CPC 15", font=f_title, border=None)
put(ws, "E2", "Hipótese A: custo original (\"magrinho\")  ·  Hipótese B: custo incrementado (\"gordinho\")", font=f_i, border=None)
put(ws, "B4", "Base: draft MLA, aba \"base 30 06 26com dividendos\" · valores em R$ milhões · células AMARELAS = premissas editáveis; o restante é fórmula.", font=f_i, border=None)
bar(ws, "B6:D6", "PREMISSAS GERAIS")
rows = [
 ("Data-base dos balanços", "30.06.2026", None, ""),
 ("Participação das Partes CNP na incorporação (relação de troca)", 0.138, PCT, "Draft E71"),
 ("Participação final das Partes CNP (pós cash out)", 0.40, PCT, "Draft E87"),
 ("IRPJ/CSLL sobre ganho de capital", 0.34, PCT, "Draft I96"),
 ("Dividendos declarados pela Embracon antes da incorporação", 120.0, NF, "Draft L23/L31 — reduz o PL da Embracon a 239,5"),
 ("Preço do cash out — por holding", 600.0, NF, "Draft G97 (total 1.200 nas duas holdings)"),
]
r = 7
for lbl, val, nf_, src in rows:
    put(ws, f"B{r}", lbl)
    put(ws, f"C{r}", val, fill=fill_yel, nf=nf_, al=(al_c if nf_ is None else None))
    put(ws, f"D{r}", src, font=f_i, al=al_l)
    if r % 2 == 1:
        ws[f"B{r}"].fill = fill_zebra; ws[f"D{r}"].fill = fill_zebra
    r += 1
bar(ws, "B14:D14", "SEGREGAÇÃO DO CUSTO — PROPORÇÃO DO EXEMPLO DELOITTE")
put(ws, "B15", "Valor justo do PL da CNP (exemplo DTT)")
put(ws, "C15", 866.0, fill=fill_yel, nf=NF)
put(ws, "D15", "Draft N86", font=f_i, al=al_l)
put(ws, "B16", "Mais-valia contida no exemplo (MV)", fill=fill_zebra)
put(ws, "C16", 300.0, fill=fill_yel, nf=NF)
put(ws, "D16", "Draft O86", font=f_i, al=al_l); ws["D16"].fill = fill_zebra
put(ws, "B17", "% AVJ implícito no PL da CNP (MV ÷ VJ)", font=f_b)
put(ws, "C17", "=C16/C15", font=f_b, nf=PCT2)
put(ws, "D17", "Aplicado ao PL contábil da CNP para segregar a camada de maior risco", font=f_i, al=al_l)
put(ws, "B18", "% contábil (sem AVJ)", font=f_b)
put(ws, "C18", "=1-C17", font=f_b, nf=PCT2)
put(ws, "D18", "", border=b_all)

P_TROCA = "Premissas!$C$8"; P_ALVO = "Premissas!$C$9"; P_IR = "Premissas!$C$10"
P_DIV = "Premissas!$C$11"; P_PRECO = "Premissas!$C$12"
P_AVJ = "Premissas!$C$17"; P_CONT = "Premissas!$C$18"

# =====================================================================
# ABA 2 — Contabilização
# =====================================================================
ws = wb.create_sheet("Contabilização")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 46, "C": 13.5, "D": 13.5, "E": 13.5, "F": 13.5})
put(ws, "B1", "Contabilização — do balanço-base à alienação futura", font=f_title, border=None)

# ---- 0. Balanços-base (rows 3-13)
bar(ws, "B3:F3", "0. BALANÇOS-BASE 30.06.2026 (R$ MILHÕES)")
hdr(ws, 4, [("C", "SAVIAN"), ("D", "JVFJ"), ("E", "EMBRACON"), ("F", "CNP")])
put(ws, "B5", "Ativos totais"); put(ws, "C5", "=C11", nf=NF); put(ws, "D5", "=D11", nf=NF)
put(ws, "E5", 26138.03, fill=fill_yel, nf=NF); put(ws, "F5", 17611.435, fill=fill_yel, nf=NF)
put(ws, "B6", "Passivos totais", fill=fill_zebra); put(ws, "C6", 0, nf=NF, fill=fill_zebra); put(ws, "D6", 0, nf=NF, fill=fill_zebra)
put(ws, "E6", 25898.882, fill=fill_yel, nf=NF); put(ws, "F6", 16771.856, fill=fill_yel, nf=NF)
put(ws, "B7", "Patrimônio líquido", font=f_b)
put(ws, "C7", "=E7/2", font=f_b, nf=NF); put(ws, "D7", "=E7/2", font=f_b, nf=NF)
put(ws, "E7", "=SUM(E8:E10)", font=f_b, nf=NF); put(ws, "F7", "=SUM(F8:F10)", font=f_b, nf=NF)
put(ws, "B8", "      Capital"); blank(ws, 8, "CD")
put(ws, "E8", 85.0, fill=fill_yel, nf=NF); put(ws, "F8", 519.0, fill=fill_yel, nf=NF)
put(ws, "B9", "      Reservas", fill=fill_zebra); blank(ws, 9, "CD"); ws["C9"].fill = fill_zebra; ws["D9"].fill = fill_zebra
put(ws, "E9", 274.488, fill=fill_yel, nf=NF); put(ws, "F9", 319.735, fill=fill_yel, nf=NF)
put(ws, "B10", "      (−) Dividendos declarados"); blank(ws, 10, "CD")
put(ws, "E10", f"=-{P_DIV}", nf=NF); put(ws, "F10", "", border=b_all)
put(ws, "B11", "Investimento na Embracon (= 50% do PL)", font=f_b)
put(ws, "C11", "=E7/2", font=f_b, nf=NF); put(ws, "D11", "=E7/2", font=f_b, nf=NF)
put(ws, "E11", "", border=b_all); put(ws, "F11", "", border=b_all)
put(ws, "B12", "PL combinado (Embracon + CNP)", font=f_b, fill=fill_zebra)
put(ws, "C12", "=E7+F7", font=f_b, nf=NF, fill=fill_zebra)
for cc in "DEF": put(ws, f"{cc}12", "", fill=fill_zebra, border=b_all)
put(ws, "B13", "Savian e JVFJ: ativo único = investimento na Embracon. CNP: PL = Capital 519 + Reservas 319,7 (draft).", font=f_i, border=None)

# ---- 1. Pré-incorporação (15-18)
bar(ws, "B15:F15", "1. PRÉ-INCORPORAÇÃO — PARTICIPAÇÕES E CUSTO")
hdr(ws, 16, [("C", "SAVIAN"), ("D", "JVFJ"), ("E", "CNP"), ("F", "TOTAL")])
put(ws, "B17", "Participação na Embracon")
put(ws, "C17", 0.50, fill=fill_yel, nf=PCT); put(ws, "D17", 0.50, fill=fill_yel, nf=PCT)
put(ws, "E17", 0, nf=PCT); put(ws, "F17", "=SUM(C17:E17)", nf=PCT)
put(ws, "B18", "Custo do investimento", fill=fill_zebra)
put(ws, "C18", "=C11", nf=NF, fill=fill_zebra); put(ws, "D18", "=D11", nf=NF, fill=fill_zebra)
put(ws, "E18", "", fill=fill_zebra, border=b_all); put(ws, "F18", "=C18+D18", nf=NF, fill=fill_zebra)

# ---- 2. Incorporação (20-41)
bar(ws, "B20:F20", "2. INCORPORAÇÃO DA CNP — RELAÇÃO DE TROCA 13,8%")
hdr(ws, 21, [("C", "SAVIAN"), ("D", "JVFJ"), ("E", "CNP"), ("F", "TOTAL")])
put(ws, "B22", "Participação após a incorporação", font=f_b)
put(ws, "C22", f"=C17*(1-{P_TROCA})", font=f_b, nf=PCT2)
put(ws, "D22", f"=D17*(1-{P_TROCA})", font=f_b, nf=PCT2)
put(ws, "E22", f"={P_TROCA}", font=f_b, nf=PCT2)
put(ws, "F22", "=SUM(C22:E22)", font=f_b, nf=PCT)
put(ws, "B24", "Reflexo nas holdings", font=f_b, fill=fill_zebra)
put(ws, "C24", "POR HOLDING", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "D24", "TOTAL (2 HOLDINGS)", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "E24", "", fill=fill_zebra, border=b_all); put(ws, "F24", "", fill=fill_zebra, border=b_all)
refl = [
 ("Investimento anterior", "=C11", "=C25*2", False),
 ("(+) Incorporação base (PL da CNP ÷ 2)", "=F7/2", "=C26*2", False),
 ("(−) Diluição de 13,8% sobre (investimento + base)", f"=-(C25+C26)*{P_TROCA}", "=C27*2", False),
 ("(=) Parcela incorporada", "=C26+C27", "=C28*2", True),
 ("(=) Custo incrementado — Hipótese B (\"gordinho\")", "=C25+C28", "=C29*2", True),
]
r = 25
for lbl, fc, fd, bold in refl:
    fnt = f_b if bold else f_t
    put(ws, f"B{r}", lbl, font=fnt)
    put(ws, f"C{r}", fc, font=fnt, nf=NF)
    put(ws, f"D{r}", fd, font=fnt, nf=NF)
    blank(ws, r, "EF")
    r += 1
put(ws, "B30", "Check: 43,1% × PL combinado − custo incrementado", font=f_i)
put(ws, "C30", "=C22*C12-C29", font=f_i, nf=NF)
blank(ws, 30, "DEF")
put(ws, "B32", "Segregação do custo (camadas de risco)", font=f_b, fill=fill_zebra)
put(ws, "C32", "PL COMBINADO", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "D32", "× 86,2%", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "E32", "Δ (2 HOLDINGS)", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "F32", "", fill=fill_zebra, border=b_all)
seg = [
 ("1. Investimento antigo (PL Embracon) — perda de diluição", "=E7", "=C33*(1-{t})", "=D33-C33"),
 ("2. Reflexo do PL contábil da CNP (menor risco)", f"=F7*{P_CONT}", "=C34*(1-{t})", "=D34"),
 ("3. Reflexo da parcela AVJ implícita (maior risco)", f"=F7*{P_AVJ}", "=C35*(1-{t})", "=D35"),
]
r = 33
for lbl, fc, fd, fe in seg:
    put(ws, f"B{r}", lbl)
    put(ws, f"C{r}", fc, nf=NF)
    put(ws, f"D{r}", fd.format(t=P_TROCA), nf=NF)
    put(ws, f"E{r}", fe, nf=NF)
    put(ws, f"F{r}", "", border=b_all)
    r += 1
put(ws, "B36", "Total", font=f_b, fill=fill_zebra)
put(ws, "C36", "=SUM(C33:C35)", font=f_b, nf=NF, fill=fill_zebra)
put(ws, "D36", "=SUM(D33:D35)", font=f_b, nf=NF, fill=fill_zebra)
put(ws, "E36", "=SUM(E33:E35)", font=f_b, nf=NF, fill=fill_zebra)
put(ws, "F36", "", fill=fill_zebra, border=b_all)
bar(ws, "B38:F38", "CONCLUSÃO DA INCORPORAÇÃO")
put(ws, "B39", "Custo original — Hipótese A (por holding)", font=f_b)
put(ws, "C39", "=C11", font=f_b, nf=NF)
blank(ws, 39, "DEF")
put(ws, "B40", "Custo incrementado — Hipótese B (por holding)", font=f_b, fill=fill_zebra)
put(ws, "C40", "=C29", font=f_b, nf=NF, fill=fill_zebra)
for cc in "DEF": put(ws, f"{cc}40", "", fill=fill_zebra, border=b_all)
put(ws, "B41", "Parcela incorporada — total (2 holdings)", font=f_b)
put(ws, "C41", "=D28", font=f_b, nf=NF)
blank(ws, 41, "DEF")

# ---- 3. SITUAÇÃO APÓS A INCORPORAÇÃO — ACERVO VERTIDO (novo, 43-88)
bar(ws, "B43:F43", "3. SITUAÇÃO APÓS A INCORPORAÇÃO — O ACERVO VERTIDO (CINDIDO) E OS BALANÇOS")

put(ws, "B44", "3.1  Acervo líquido da CNP vertido à Embracon (linha a linha)", font=f_b, fill=fill_zebra)
put(ws, "C44", "VALOR", font=f_b, fill=fill_zebra, al=al_c)
for cc in "DEF": put(ws, f"{cc}44", "", fill=fill_zebra, border=b_all)
acervo = [
 ("Ativos — Circulante + realizável LP", 945.234, True),
 ("Ativos — CPC 47 (despesas/comissões diferidas)", 216.353, True),
 ("Ativos — Permanente", 138.693, True),
 ("Ativos — Compensação (grupos de consórcio)", 16311.155, True),
]
r = 45
for lbl, val, is_input in acervo:
    put(ws, f"B{r}", lbl)
    put(ws, f"C{r}", val, fill=fill_yel, nf=NF)
    blank(ws, r, "DEF")
    r += 1
put(ws, "B49", "Total de ativos vertidos", font=f_b)
put(ws, "C49", "=SUM(C45:C48)", font=f_b, nf=NF)
blank(ws, 49, "DEF")
put(ws, "B50", "(−) Passivos — Circulante")
put(ws, "C50", -232.757, fill=fill_yel, nf=NF)
blank(ws, 50, "DEF")
put(ws, "B51", "(−) Passivos — CPC 47 (receitas diferidas)")
put(ws, "C51", -227.944, fill=fill_yel, nf=NF)
blank(ws, 51, "DEF")
put(ws, "B52", "(−) Passivos — Compensação (grupos de consórcio)")
put(ws, "C52", "=-C48", nf=NF)
blank(ws, 52, "DEF")
put(ws, "B53", "Total de passivos assumidos", font=f_b)
put(ws, "C53", "=SUM(C50:C52)", font=f_b, nf=NF)
blank(ws, 53, "DEF")
put(ws, "B54", "ACERVO LÍQUIDO VERTIDO", font=f_b, fill=fill_zebra)
put(ws, "C54", "=C49+C53", font=f_b, nf=NF, fill=fill_zebra)
for cc in "DEF": put(ws, f"{cc}54", "", fill=fill_zebra, border=b_all)
put(ws, "B55", "Check: acervo líquido − PL contábil da CNP (diferença da base do draft, a conciliar)", font=f_i)
put(ws, "C55", "=C54-F7", font=f_i, nf=NF)
blank(ws, 55, "DEF")

put(ws, "B57", "3.2  Embracon combinada — balanço após a incorporação", font=f_b, fill=fill_zebra)
put(ws, "C57", "EMBRACON ANTES", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "D57", "ACERVO VERTIDO (CNP)", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "E57", "EMBRACON APÓS", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "F57", "", fill=fill_zebra, border=b_all)
comb = [
 ("Ativo — Circulante", 328.578, "=C45"),
 ("Ativo — CPC 47 (despesas/comissões diferidas)", 1937.402, "=C46"),
 ("Ativo — Permanente", 57.986, "=C47"),
 ("Ativo — Compensação", 23814.064, "=C48"),
]
r = 58
for lbl, antes, vert in comb:
    put(ws, f"B{r}", lbl)
    put(ws, f"C{r}", antes, fill=fill_yel, nf=NF)
    put(ws, f"D{r}", vert, nf=NF)
    put(ws, f"E{r}", f"=C{r}+D{r}", nf=NF)
    put(ws, f"F{r}", "", border=b_all)
    r += 1
put(ws, "B62", "Total de ativos", font=f_b)
put(ws, "C62", "=SUM(C58:C61)", font=f_b, nf=NF)
put(ws, "D62", "=SUM(D58:D61)", font=f_b, nf=NF)
put(ws, "E62", "=C62+D62", font=f_b, nf=NF)
put(ws, "F62", "", border=b_all)
passv = [
 ("Passivo — Circulante", 420.060, "=-C50"),
 ("Passivo — CPC 47 (receitas diferidas)", 1544.758, "=-C51"),
]
r = 63
for lbl, antes, vert in passv:
    put(ws, f"B{r}", lbl)
    put(ws, f"C{r}", antes, fill=fill_yel, nf=NF)
    put(ws, f"D{r}", vert, nf=NF)
    put(ws, f"E{r}", f"=C{r}+D{r}", nf=NF)
    put(ws, f"F{r}", "", border=b_all)
    r += 1
put(ws, "B65", "Passivo — Dividendos a pagar")
put(ws, "C65", f"={P_DIV}", nf=NF)
put(ws, "D65", "", border=b_all)
put(ws, "E65", "=C65", nf=NF)
put(ws, "F65", "", border=b_all)
put(ws, "B66", "Passivo — Compensação")
put(ws, "C66", "=C61", nf=NF)
put(ws, "D66", "=D61", nf=NF)
put(ws, "E66", "=C66+D66", nf=NF)
put(ws, "F66", "", border=b_all)
put(ws, "B67", "Total de passivos", font=f_b)
put(ws, "C67", "=SUM(C63:C66)", font=f_b, nf=NF)
put(ws, "D67", "=SUM(D63:D66)", font=f_b, nf=NF)
put(ws, "E67", "=C67+D67", font=f_b, nf=NF)
put(ws, "F67", "", border=b_all)
plrows = [
 ("PL — Capital social", "=E8", "=F8"),
 ("PL — Reservas", "=E9", "=F9"),
 ("PL — (−) Dividendos declarados", "=E10", ""),
]
r = 68
for lbl, antes, vert in plrows:
    put(ws, f"B{r}", lbl)
    put(ws, f"C{r}", antes, nf=NF)
    if vert:
        put(ws, f"D{r}", vert, nf=NF)
    else:
        put(ws, f"D{r}", "", border=b_all)
    put(ws, f"E{r}", f"=C{r}+D{r}" if vert else f"=C{r}", nf=NF)
    put(ws, f"F{r}", "", border=b_all)
    r += 1
put(ws, "B71", "Patrimônio líquido após", font=f_b, fill=fill_zebra)
put(ws, "C71", "=SUM(C68:C70)", font=f_b, nf=NF, fill=fill_zebra)
put(ws, "D71", "=SUM(D68:D70)", font=f_b, nf=NF, fill=fill_zebra)
put(ws, "E71", "=C71+D71", font=f_b, nf=NF, fill=fill_zebra)
put(ws, "F71", "", fill=fill_zebra, border=b_all)
put(ws, "B72", "Check (Ativos − Passivos − PL): diferenças da base do draft, a conciliar", font=f_i)
put(ws, "C72", "=C62-C67-C71", font=f_i, nf=NF)
put(ws, "D72", "=D62-D67-D71", font=f_i, nf=NF)
put(ws, "E72", "=E62-E67-E71", font=f_i, nf=NF)
put(ws, "F72", "", border=b_all)

put(ws, "B74", "3.3  Balanço das holdings após (abertura do investimento)", font=f_b, fill=fill_zebra)
put(ws, "C74", "SAVIAN", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "D74", "JVFJ", font=f_b, fill=fill_zebra, al=al_c)
for cc in "EF": put(ws, f"{cc}74", "", fill=fill_zebra, border=b_all)
h3 = [
 ("Inv. Embracon — custo antigo", "=C11", "=D11", False),
 ("(+) Incorporação base (acervo ÷ 2, a valor contábil)", "=C26", "=C26", False),
 ("(−) Diluição de 13,8%", "=C27", "=C27", False),
 ("(=) Investimento e PL da holding após", "=SUM(C75:C77)", "=SUM(D75:D77)", True),
]
r = 75
for lbl, fc, fd, bold in h3:
    fnt = f_b if bold else f_t
    put(ws, f"B{r}", lbl, font=fnt)
    put(ws, f"C{r}", fc, font=fnt, nf=NF)
    put(ws, f"D{r}", fd, font=fnt, nf=NF)
    blank(ws, r, "EF")
    r += 1
put(ws, "B79", "Check: 43,1% × PL combinado após − investimento da holding", font=f_i)
put(ws, "C79", "=C22*E71-C78", font=f_i, nf=NF)
blank(ws, 79, "DEF")

put(ws, "B81", "3.4  Prova da somatória (como no draft)", font=f_b, fill=fill_zebra)
put(ws, "C81", "VALOR", font=f_b, fill=fill_zebra, al=al_c)
for cc in "DEF": put(ws, f"{cc}81", "", fill=fill_zebra, border=b_all)
prova = [
 ("PL combinado após a incorporação", "=E71", False),
 ("Participação das Partes CNP (13,8%)", f"=E71*{P_TROCA}", False),
 ("Participação Savian + JVFJ (86,2%)", f"=E71*(1-{P_TROCA})", False),
 ("      por holding (÷ 2)", "=C84/2", True),
 ("Diluição individual (participação CNP ÷ 2)", "=C83/2", False),
]
r = 82
for lbl, fc, bold in prova:
    fnt = f_b if bold else f_t
    put(ws, f"B{r}", lbl, font=fnt)
    put(ws, f"C{r}", fc, font=fnt, nf=NF)
    blank(ws, r, "DEF")
    r += 1
put(ws, "B87", "Check: inv. anterior (239,5) + parcela incorporada (690,0) − participação S+J", font=f_i)
put(ws, "C87", "=E7+D28-C84", font=f_i, nf=NF)
blank(ws, 87, "DEF")
put(ws, "B88", "CNP Consórcio: extinta na incorporação (balanço zerado). O acervo entra a valor contábil; o risco fiscal está no reflexo que ele gera nas holdings — segregado em 3 camadas na seção 2 (a camada AVJ é a de maior risco).", font=f_i, border=None)

# ---- 4. Cash out (90-109)
bar(ws, "B90:F90", "4. CASH OUT — VENDA DE 26,2% À CNP (600 POR HOLDING)")
hdr(ws, 91, [("C", "SAVIAN"), ("D", "JVFJ"), ("E", "CNP"), ("F", "TOTAL")])
put(ws, "B92", "Participação após a incorporação")
put(ws, "C92", "=C22", nf=PCT2); put(ws, "D92", "=D22", nf=PCT2); put(ws, "E92", "=E22", nf=PCT2); put(ws, "F92", "=SUM(C92:E92)", nf=PCT)
put(ws, "B93", "Venda / compra", fill=fill_zebra)
put(ws, "C93", f"=-({P_ALVO}-{P_TROCA})/2", nf=PCT2, fill=fill_zebra)
put(ws, "D93", "=C93", nf=PCT2, fill=fill_zebra)
put(ws, "E93", f"={P_ALVO}-{P_TROCA}", nf=PCT2, fill=fill_zebra)
put(ws, "F93", "=SUM(C93:E93)", nf=PCT, fill=fill_zebra)
put(ws, "B94", "Participação final", font=f_b)
put(ws, "C94", "=C92+C93", font=f_b, nf=PCT); put(ws, "D94", "=D92+D93", font=f_b, nf=PCT)
put(ws, "E94", "=E92+E93", font=f_b, nf=PCT); put(ws, "F94", "=SUM(C94:E94)", font=f_b, nf=PCT)
put(ws, "B95", "Fração alienada do investimento (13,1 ÷ 43,1)", font=f_b)
put(ws, "C95", "=-C93/C92", font=f_b, nf=PCT2)
blank(ws, 95, "DEF")
put(ws, "B97", "Ganho de capital no cash out (por holding)", font=f_b, fill=fill_zebra)
put(ws, "C97", "HIP. A — CUSTO ORIGINAL", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "D97", "HIP. B — CUSTO INCREMENTADO", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "E97", "", fill=fill_zebra, border=b_all); put(ws, "F97", "", fill=fill_zebra, border=b_all)
gk = [
 ("Preço da venda", f"={P_PRECO}", f"={P_PRECO}", False),
 ("(−) Custo alocado (custo × fração alienada)", "=-C11*$C$95", "=-C29*$C$95", False),
 ("(=) Ganho de capital", "=C98+C99", "=D98+D99", False),
 ("IRPJ/CSLL (34%)", f"=C100*{P_IR}", f"=D100*{P_IR}", True),
 ("Líquido do imposto (preço − IR)", "=C98-C101", "=D98-D101", False),
 ("Custo remanescente (para a alienação futura)", "=C11+C99", "=C29+D99", False),
]
r = 98
for lbl, fa, fb_, bold in gk:
    fnt = f_b if bold else f_t
    put(ws, f"B{r}", lbl, font=fnt)
    put(ws, f"C{r}", fa, font=fnt, nf=NF)
    put(ws, f"D{r}", fb_, font=fnt, nf=NF)
    blank(ws, r, "EF")
    r += 1
bar(ws, "B105:F105", "CONCLUSÃO DO CASH OUT (TOTAL — 2 HOLDINGS)")
put(ws, "B106", "IRPJ/CSLL — Hipótese A", font=f_b)
put(ws, "C106", "=C101*2", font=f_b, nf=NF)
blank(ws, 106, "DEF")
put(ws, "B107", "IRPJ/CSLL — Hipótese B", font=f_b, fill=fill_zebra)
put(ws, "C107", "=D101*2", font=f_b, nf=NF, fill=fill_zebra)
for cc in "DEF": put(ws, f"{cc}107", "", fill=fill_zebra, border=b_all)
put(ws, "B108", "Economia da Hipótese B no cash out", font=f_b)
put(ws, "C108", "=C106-C107", font=f_b, nf=NF)
blank(ws, 108, "DEF")
put(ws, "B109", "Operação única (primária + secundária simultâneas): no cash out prevalece o custo antigo — a Hipótese A é a leitura conservadora.", font=f_i, border=None)

# ---- 5. Alienação futura (111-121)
bar(ws, "B111:F111", "5. ALIENAÇÃO FUTURA — VENDA DO REMANESCENTE (30% POR HOLDING)")
put(ws, "B112", "Preço da venda futura — por holding (editável)", font=f_b)
put(ws, "C112", 600.0, font=f_b, fill=fill_yel, nf=NF)
blank(ws, 112, "DEF")
put(ws, "B113", "", font=f_b, fill=fill_zebra, border=b_all)
put(ws, "C113", "HIP. A — CUSTO ORIGINAL", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "D113", "HIP. B — CUSTO INCREMENTADO", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "E113", "", fill=fill_zebra, border=b_all); put(ws, "F113", "", fill=fill_zebra, border=b_all)
af = [
 ("Custo remanescente", "=C103", "=D103", False),
 ("Ganho de capital (preço − custo)", "=$C$112-C114", "=$C$112-D114", False),
 ("IRPJ/CSLL (34%; não negativo)", f"=MAX(0,C115)*{P_IR}", f"=MAX(0,D115)*{P_IR}", True),
]
r = 114
for lbl, fa, fb_, bold in af:
    fnt = f_b if bold else f_t
    put(ws, f"B{r}", lbl, font=fnt)
    put(ws, f"C{r}", fa, font=fnt, nf=NF)
    put(ws, f"D{r}", fb_, font=fnt, nf=NF)
    blank(ws, r, "EF")
    r += 1
bar(ws, "B118:F118", "CONCLUSÃO DA ALIENAÇÃO FUTURA (TOTAL — 2 HOLDINGS)")
put(ws, "B119", "Economia de IR da Hipótese B no preço indicado", font=f_b)
put(ws, "C119", "=(C116-D116)*2", font=f_b, nf=NF)
blank(ws, 119, "DEF")
put(ws, "B120", "Economia máxima (34% × Δ de custo remanescente)", font=f_b, fill=fill_zebra)
put(ws, "C120", f"=(D114-C114)*2*{P_IR}", font=f_b, nf=NF, fill=fill_zebra)
for cc in "DEF": put(ws, f"{cc}120", "", fill=fill_zebra, border=b_all)
put(ws, "B121", "A economia independe do preço enquanto o ganho for positivo nas duas hipóteses.", font=f_i, border=None)

INC = "Contabilização"

# =====================================================================
# ABA 3 — Conclusão
# =====================================================================
ws = wb.create_sheet("Conclusão")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 50, "C": 15, "D": 15, "E": 15})
put(ws, "B1", "Conclusão — efeitos das duas hipóteses de custo", font=f_title, border=None)
bar(ws, "B3:E3", "SÍNTESE (TOTAL DAS 2 HOLDINGS — R$ MILHÕES)")
put(ws, "B4", "", font=f_b, fill=fill_zebra, border=b_all)
put(ws, "C4", "HIP. A — CUSTO ORIGINAL", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "D4", "HIP. B — CUSTO INCREMENTADO", font=f_b, fill=fill_zebra, al=al_c)
put(ws, "E4", "Δ (ECONOMIA B)", font=f_b, fill=fill_zebra, al=al_c)
lin = [
 ("Custo do investimento após a incorporação", f"={INC}!C11*2", f"={INC}!C29*2", "=D5-C5", False),
 ("IRPJ/CSLL no cash out", f"={INC}!C101*2", f"={INC}!D101*2", "=C6-D6", True),
 ("Custo remanescente para a alienação futura", f"={INC}!C103*2", f"={INC}!D103*2", "=D7-C7", False),
 ("IRPJ/CSLL na alienação futura (preço editável da aba anterior)", f"={INC}!C116*2", f"={INC}!D116*2", "=C8-D8", True),
 ("IRPJ/CSLL total (cash out + alienação futura)", "=C6+C8", "=D6+D8", "=C9-D9", True),
]
r = 5
for lbl, fa, fb_, fe, bold in lin:
    fnt = f_b if bold else f_t
    put(ws, f"B{r}", lbl, font=fnt)
    put(ws, f"C{r}", fa, font=fnt, nf=NF)
    put(ws, f"D{r}", fb_, font=fnt, nf=NF)
    put(ws, f"E{r}", fe, font=fnt, nf=NF)
    if not bold:
        for cc in "BCDE": ws[f"{cc}{r}"].fill = fill_zebra
    r += 1
put(ws, "B10", "Benefício máximo da tese do custo \"gordinho\" (34% × parcela incorporada)", font=f_b, fill=fill_zebra)
put(ws, "C10", "", fill=fill_zebra, border=b_all)
put(ws, "D10", "", fill=fill_zebra, border=b_all)
put(ws, "E10", f"={INC}!D28*{P_IR}", font=f_b, nf=NF, fill=fill_zebra)
put(ws, "B11", "Check: economia no cash out + economia máxima na alienação − benefício máximo", font=f_i)
put(ws, "C11", "", border=b_all); put(ws, "D11", "", border=b_all)
put(ws, "E11", f"=E6+{INC}!C120-E10", font=f_i, nf=NF)

bar(ws, "B13:E13", "LEITURA")
leitura = [
 ("Onde nasce a diferença", "Parcela incorporada de 345,0 por holding (690,0 no total) = reflexo do PL contábil da CNP (419,4) − diluição de 13,8% (74,4). Compõe custo na Hipótese B; não compõe na Hipótese A."),
 ("Risco por camada", "Perda de diluição (33,0): menor controvérsia · reflexo do PL contábil da CNP (472,5): neutralidade defensável (MEP / variação de participação) · reflexo da parcela AVJ implícita (250,5): maior risco de requalificação."),
 ("Enquadramento", "A definição CPC 15 x CPC 19 cabe ao auditor. Sob CPC 19, o AVJ fica \"carimbado\" a 34% em subconta e a Hipótese B deixa de existir."),
]
r = 14
for t, d in leitura:
    put(ws, f"B{r}", t, font=f_b, al=Alignment(horizontal="left", vertical="top", wrap_text=True))
    ws.merge_cells(f"C{r}:E{r}")
    put(ws, f"C{r}", d, al=Alignment(horizontal="left", vertical="top", wrap_text=True))
    for cc in "DE": ws[f"{cc}{r}"].border = b_all
    ws.row_dimensions[r].height = 30
    r += 1

wb.save(OUT)
print("saved", OUT)
