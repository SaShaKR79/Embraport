import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "uploads/user_cpc15_v1.xlsx"
OUT = "Exercicio_Incorporacao_CPC15_v1.1.xlsx"
wb = openpyxl.load_workbook(SRC)
idx = wb.sheetnames.index("Conclusão")
ws = wb.create_sheet("6. Segregação de Risco", idx)
ws.sheet_view.showGridLines = False

RED="FFC00000"; ZEB="FFEAF0FB"; GRAY="FFC9C9C9"
f_title=Font(name="Calibri",size=12,bold=True); f_bar=Font(name="Calibri",size=9,bold=True,color="FFFFFFFF")
f_b=Font(name="Calibri",size=9,bold=True); f_t=Font(name="Calibri",size=9); f_i=Font(name="Calibri",size=9,italic=True)
f_ig=Font(name="Calibri",size=9,italic=True,color="FF7F7F7F"); f_in=Font(name="Calibri",size=9,color="FF0000FF"); f_inb=Font(name="Calibri",size=9,bold=True,color="FF0000FF")
fill_red=PatternFill("solid",fgColor=RED); fill_z=PatternFill("solid",fgColor=ZEB); fill_y=PatternFill("solid",fgColor="FFFFF2CC")
thin=Side(style="thin",color=GRAY); B=Border(left=thin,right=thin,top=thin,bottom=thin)
al_c=Alignment(horizontal="center",vertical="center",wrap_text=True); al_l=Alignment(horizontal="left",vertical="center",wrap_text=True); al_r=Alignment(horizontal="right",vertical="center")
NF="#,##0.0;[Red]\\(#,##0.0\\)"; PCT="0.0%"; PCT2="0.00%"
for col,w in {"A":2.5,"B":52,"C":15,"D":15,"E":15,"F":15,"G":15,"H":58}.items(): ws.column_dimensions[col].width=w

def put(coord,v,font=f_t,fill=None,nf=None,al=None,border=B):
    c=ws[coord]; c.value=v; c.font=font
    if fill: c.fill=fill
    if nf: c.number_format=nf
    c.alignment=al or (al_r if nf else al_l)
    if border: c.border=border
    return c
def bar(rng,text):
    ws.merge_cells(rng); a,b=rng.split(":"); row=int(''.join(ch for ch in a if ch.isdigit()))
    c1=openpyxl.utils.column_index_from_string(''.join(ch for ch in a if ch.isalpha())); c2=openpyxl.utils.column_index_from_string(''.join(ch for ch in b if ch.isalpha()))
    put(a,text,font=f_bar,fill=fill_red,al=al_c)
    for cc in range(c1,c2+1): ws.cell(row=row,column=cc).border=B; ws.cell(row=row,column=cc).fill=fill_red
def hdr(row,labels,first=""):
    put(f"B{row}",first,font=f_b,fill=fill_z)
    for col,l in labels: put(f"{col}{row}",l,font=f_b,fill=fill_z,al=al_c)
def blank(row,cols):
    for col in cols: put(f"{col}{row}","",border=B)
def note(row,text,cols="B:H"):
    ws.merge_cells(f"{cols[0]}{row}:{cols[2]}{row}"); put(f"{cols[0]}{row}",text,font=f_i,border=None); ws.row_dimensions[row].height=26

SA="'1. Situação Antes'"; P="Premissas"
put("B1","6. Segregação de risco do custo do investimento — CPC 15 × CPC 19 e venda futura dos 60% (reprodução dos quadros da MLA)",font=f_title,border=None)

# ---------- Premissas
bar("B3:H3","PREMISSAS DESTA ABA (VINCULADAS ÀS DEMAIS ABAS; AZUL = INPUT DO EXERCÍCIO DTT/MLA)")
prem=[
 ("PL da Embracon 30.06.26 (com dividendos)",f"={SA}!I18",NF,"aba 1"),
 ("PL da CNP 30.06.26",f"={SA}!L18",NF,"aba 1"),
 ("PL combinado (Embracon + CNP)","=C4+C5",NF,""),
 ("Participação das Partes CNP na incorporação",f"={P}!C6",PCT,"Premissas"),
 ("Participação final das Partes CNP (pós cash out)",f"={P}!C7",PCT,"Premissas"),
 ("Participação final das holdings (posição remanescente)","=1-C8",PCT,""),
 ("IRPJ/CSLL sobre ganho de capital (PJ)",f"={P}!C10",PCT,"Premissas"),
 ("% contábil do PL da CNP (proporção DTT: 566 ÷ 866)",f"={P}!C15",PCT2,"Premissas"),
 ("% AVJ implícito no PL da CNP (proporção DTT: 300 ÷ 866)",f"={P}!C14",PCT2,"Premissas"),
 ("VJ da JO — 100% (exercício DTT)",5888.0,NF,"draft O90 — input"),
 ("VJ da CNP (exercício DTT)",f"={P}!C12",NF,"Premissas (866)"),
 ("VJ da Embracon (residual = VJ JO − VJ CNP)","=C13-C14",NF,"draft O92"),
 ("Custo antigo baixado na secundária — critério do draft (13,1% × custo, 2 holdings)","=C4*(C8-C7)/2",NF,"draft F100: 239,5 × 13,1% = 31,4 (a aba 3 usa rateio 13,1/43,1)"),
 ("Custo do investimento das PFs nas holdings (draft: 18)",18.0,NF,"draft Z78 — input"),
 ("Alíquota simplificada de IRPF sobre o ganho (venda pela PF)",0.225,PCT,"draft — input"),
 ("Preço de venda futura dos 60% — cenário (i) inferior ao EV",1000.0,NF,"input"),
 ("Preço de venda futura dos 60% — cenário (ii) = VJ da posição (60% × VJ JO)","=C13*C9",NF,"3.532,8"),
 ("Preço de venda futura dos 60% — cenário (iii) superior ao EV",5000.0,NF,"input"),
]
r=4
for lbl,v,nf,src in prem:
    is_input = not (isinstance(v,str) and v.startswith("="))
    put(f"B{r}",lbl); put(f"C{r}",v,font=(f_in if is_input else f_t),fill=(fill_y if is_input else None),nf=nf)
    blank(r,"DEFG"); put(f"H{r}",src,font=f_ig,al=al_l)
    if r%2==0:
        for cc in "BDEFGH": ws[f"{cc}{r}"].fill=fill_z
    r+=1
# fix: IR na Premissas do usuário está em C8 (não C10)
ws["C10"]=f"={P}!C8"

# ---------- A. CPC 15
bar("B23:H23","A. CPC 15 — SEGREGAÇÃO DE RISCO DO CUSTO DO INVESTIMENTO (POSIÇÃO FINAL DE 60%) — QUADRO 'MLA - SEGREGAÇÃO RISCO CUSTO DO INVESTIMENTO'")
hdr(24,[("C","TOTAL (PL combinado)"),("D","ANTES"),("E","DEPOIS (60%)"),("F","Δ"),("H","LEITURA")],"Camada do custo")
put("G24","",fill=fill_z,border=B)
rowsA=[
 ("1. Inv Embracon — custo antigo","=C4","=C4","=D25-C16","=E25-D25","Custo histórico das holdings menos o custo baixado na secundária (31,4). No draft, rotulado 'PERDA DILUIÇÃO'. Camada incontroversa."),
 ("2. Inv CNP (PL contábil) — reflexo PL","=C5*C11",0,"=E28-E25-E27","=E26-D26","Reflexo do PL contábil da CNP: residual para fechar 60% × PL combinado (draft T85 = 646,9 − 208,1 − 174,3). Neutralidade defensável (MEP / variação de participação) — risco menor."),
 ("3. Inv CNP (AVJ) — reflexo AVJ","=C5*C12",0,"=C27*C9","=E27-D27","Reflexo da parcela AVJ implícita no PL da CNP (34,6%), proporcional à posição (60% × 290,6). Maior risco de requalificação (AVJ reflexo)."),
]
r=25
for lbl,ct,an,de,dl,le in rowsA:
    put(f"B{r}",lbl); put(f"C{r}",ct,nf=NF); put(f"D{r}",an,nf=NF); put(f"E{r}",de,nf=NF); put(f"F{r}",dl,nf=NF); put(f"G{r}","",border=B); put(f"H{r}",le,font=f_i,al=al_l)
    r+=1
put("B28","NOVO EMBRACON — custo do investimento (Hipótese B, 'gordinho')",font=f_b,fill=fill_z)
put("C28","=SUM(C25:C27)",font=f_b,nf=NF,fill=fill_z); put("D28","=SUM(D25:D27)",font=f_b,nf=NF,fill=fill_z)
put("E28","=C28*C9",font=f_b,nf=NF,fill=fill_z); put("F28","=E28-D28",font=f_b,nf=NF,fill=fill_z); put("G28","",fill=fill_z,border=B)
put("H28","DEPOIS = 60% × PL combinado (1.078,2 × 60% = 646,9): é o custo incrementado da posição remanescente — o mesmo 'custo remanescente B' da aba 3, na visão do draft.",font=f_i,fill=fill_z,al=al_l)
put("B29","Check: soma das camadas − 60% × PL combinado",font=f_ig); put("C29","=E25+E26+E27-E28",font=f_ig,nf=NF); blank(29,"DEFG"); put("H29","",border=B)
put("B30","Memo (draft S85): Inv Embracon depois + AVJ depois",font=f_ig); put("C30","=E25+E27",font=f_ig,nf=NF); blank(30,"DEFG"); put("H30","382,4 — as duas camadas 'diretas'; o reflexo PL é o que sobra (T85 = 264,5).",font=f_ig,al=al_l)
put("B31","Memo: alíquota × camadas de reflexo (IR potencial se negadas): reflexo PL · reflexo AVJ",font=f_ig); put("C31","=E26*C10",font=f_ig,nf=NF); put("D31","=E27*C10",font=f_ig,nf=NF); blank(31,"EFG"); put("H31","89,9 e 59,3 — são exatamente os 'degraus' dos riscos 1 e 2 do bloco D.",font=f_ig,al=al_l)

# ---------- B. CPC 19
bar("B33:H33","B. CPC 19 — SEGREGAÇÃO DO AVJ REFLEXO NAS HOLDINGS (JOINT OPERATION) — QUADRO 'MLA - SEGREGAÇÃO AVJ REFLEXO HOLDINGS DA EMBRACON'")
hdr(34,[("C","100% JO"),("D","60% (holdings)"),("H","LEITURA")],"Item")
for cc in "EFG": put(f"{cc}34","",fill=fill_z,border=B)
rowsB=[
 ("Inv JO (contábil) = PL Embracon + parcela contábil do PL da CNP","=C4+C5*C11","=C35*C9","Base contábil da JO (239,5 + 548,2 = 787,7); 60% = 472,6. É o custo fiscal 'magro' na JO — reflexo patrimonial."),
 ("AVJ JO = VJ da JO − base contábil","=C13-C35","=C13*C9-D35","Reavaliação a valor justo de TODO o negócio (CNP e Embracon): 5.100,3; 60% = 3.060,2. Diferido em subconta; passivo fiscal diferido."),
 ("NOVO EMBRACON a valor justo (Inv JO)","=C35+C36","=D35+D36","5.888 (100%) e 3.532,8 (60%) — valor pelo qual as holdings reconhecem a participação no CPC 19."),
 ("'Carimbo': tributação diferida do AVJ (34%)","=C36*C10","=D36*C10","1.040,5 nos 60% — devido na realização (alienação/baixa), qualquer que seja o preço; é o piso da carga no CPC 19."),
]
r=35
for lbl,c100,c60,le in rowsB:
    bold = lbl.startswith(("NOVO","'Carimbo"))
    fnt=f_b if bold else f_t
    put(f"B{r}",lbl,font=fnt); put(f"C{r}",c100,font=fnt,nf=NF); put(f"D{r}",c60,font=fnt,nf=NF); blank(r,"EFG"); put(f"H{r}",le,font=f_i,al=al_l)
    if bold:
        for cc in "BCDEFGH": ws[f"{cc}{r}"].fill=fill_z
    r+=1

# ---------- C. Venda futura
bar("B40:H40","C. VENDA FUTURA DOS 60% — PJ × PF × CPC 15 / CPC 19 — QUADRO 'VENDA PJ / VENDA PF'")
hdr(41,[("C","CPC 15 — venda pela PJ"),("D","CPC 19 — venda pela PJ"),("E","CPC 15 — venda pela PF"),("F","CPC 19 — venda pela PF"),("H","LEITURA")],"Preço da venda (60%)")
put("G41","",fill=fill_z,border=B)
leC=["Preço inferior ao EV: no CPC 19 o carimbo (1.040) prevalece sobre 34% × ganho; no CPC 15 tributa-se só o ganho sobre o custo 'gordinho' (646,9).",
     "Preço = VJ da posição: CPC 15 981 × CPC 19 1.040 — a diferença (59,3) é 34% × camada AVJ (174,3): no CPC 19 o AVJ da própria Embracon é tributado; no CPC 15 não é reavaliado.",
     "Preço superior ao EV: CPC 19 = 34% × (preço − 472,6) — tributa integralmente o excedente sobre a base contábil; PF paga o carimbo na holding + 22,5% sobre o excedente acima do VJ."]
for i,pr in enumerate(["=C19","=C20","=C21"]):
    r=42+i
    put(f"B{r}",pr,font=f_b,nf=NF)
    put(f"C{r}",f"=(B{r}-$E$28)*$C$10",nf=NF)
    put(f"D{r}",f"=MAX($D$38,(B{r}-$D$35)*$C$10)",nf=NF)
    put(f"E{r}",f"=(B{r}-$C$17)*$C$18",nf=NF)
    put(f"F{r}",f"=$D$38+MAX(0,B{r}-$C$20)*$C$18",nf=NF)
    put(f"G{r}","",border=B); put(f"H{r}",leC[i],font=f_i,al=al_l)
note(45,"Fórmulas: CPC 15 PJ = (preço − custo gordinho 646,9) × 34% · CPC 19 PJ = máximo entre o carimbo (1.040,5) e (preço − base contábil 472,6) × 34% · CPC 15 PF = (preço − custo PF 18) × 22,5% · CPC 19 PF = carimbo + 22,5% × excedente sobre o VJ. No draft, os valores de CPC 19 a 1.000 e a 3.533 estão 'travados' no carimbo — aqui generalizado pela função MÁXIMO.")

# ---------- D. Riscos
bar("B47:H47","D. CPC 15 — TESTE DE RISCO: SE O FISCO NEGAR AS CAMADAS DE REFLEXO (VENDA PELA PJ) — QUADROS 'VENDA PJ - RISCO 2' E 'RISCO 1'")
hdr(48,[("C","Custo gordinho (3 camadas)"),("D","RISCO 2 — nega a camada AVJ"),("E","RISCO 1 — nega AVJ e PL"),("F","CPC 19 (referência)"),("H","LEITURA")],"Preço da venda (60%)")
put("G48","",fill=fill_z,border=B)
leD=["Custo aceito: 646,9 → 472,6 (risco 2) → 208,1 (risco 1). Cada degrau de IR = 34% × camada negada (59,3 pela AVJ; 89,9 pelo reflexo PL).",
     "A preço = VJ: 981 → 1.040 → 1.130. Note que RISCO 2 (CPC 15 sem a camada AVJ) coincide com o CPC 19: ambos partem do custo contábil de 472,6.",
     "A preço superior: os degraus se mantêm (59,3 e 89,9) — o risco é fixo em valor, independe do preço."]
for i in range(3):
    r=49+i
    put(f"B{r}",f"=B{42+i}",font=f_b,nf=NF)
    put(f"C{r}",f"=(B{r}-$E$28)*$C$10",nf=NF)
    put(f"D{r}",f"=(B{r}-($E$25+$E$26))*$C$10",nf=NF)
    put(f"E{r}",f"=(B{r}-$E$25)*$C$10",nf=NF)
    put(f"F{r}",f"=D{42+i}",nf=NF)
    put(f"G{r}","",border=B); put(f"H{r}",leD[i],font=f_i,al=al_l)

# ---------- E. Conferência CPC 19
bar("B54:H54","E. CPC 19 — CONFERÊNCIA DA CARGA NA VENDA A PREÇO INFERIOR AO VJ (QUADRO AUXILIAR 'BAIXA / ENTRA / RESULTADO' DO DRAFT)")
hdr(55,[("C","Valor"),("H","LEITURA")],"Item (preço do cenário i)")
for cc in "DEFG": put(f"{cc}55","",fill=fill_z,border=B)
rowsE=[
 ("Baixa da base contábil da JO (60%)","=-D35","Baixa do custo contábil (472,6)."),
 ("Baixa do AVJ reconhecido (60%)","=-D36","Baixa da reavaliação (3.060,2) — o ativo estava a VJ."),
 ("Entrada: preço de venda","=C19","Cenário (i)."),
 ("Resultado contábil da venda","=SUM(C56:C58)","Prejuízo contábil de 2.532,8 ao vender abaixo do VJ."),
 ("IR/CS sobre o resultado contábil (34%)","=C59*C10","Crédito teórico de 861,2 — só se o prejuízo fosse integralmente aproveitável."),
 ("(+) Tributação do AVJ carimbado (34% × 3.060,2)","=D38","O ganho diferido é realizado na alienação: 1.040,5."),
 ("Carga líquida teórica (se o prejuízo compensasse)","=C60+C61","179,3 = 34% × (1.000 − 472,6): equivale a tributar só o ganho econômico."),
 ("Memo: 34% × (preço − base contábil)","=(C58-D35)*C10","Conferência: igual à linha anterior (draft X98 'ir/cs gk')."),
 ("Carga adotada na tabela (piso = carimbo)","=D38","Conservador: a perda na venda abaixo do VJ não compensa de imediato (prejuízo fiscal limitado a 30%/ano) — por isso o draft trava 1.040 nos preços ≤ VJ."),
 ("Memo (draft X100): ganho × 34% + carimbo, sem compensar a perda","=C63+D38","1.219,8 — leitura mais gravosa, se o AVJ e o ganho fossem tributados separadamente sem aproveitar a perda."),
]
r=56
for lbl,v,le in rowsE:
    bold = lbl.startswith(("Carga","Resultado"))
    fnt=f_b if bold else f_t
    put(f"B{r}",lbl,font=fnt); put(f"C{r}",v,font=fnt,nf=NF); blank(r,"DEFG"); put(f"H{r}",le,font=f_i,al=al_l)
    if bold:
        for cc in "BCDEFGH": ws[f"{cc}{r}"].fill=fill_z
    r+=1

# ---------- F. Leitura
bar("B67:H67","F. LEITURA — O QUE OS QUADROS DA MLA DEMONSTRAM")
leit=[
 ("CPC 15 (custo)","O custo 'gordinho' da posição de 60% (646,9) não é homogêneo: 208,1 é custo histórico incontroverso; 264,5 é reflexo do PL contábil da CNP (tese de neutralidade do MEP/variação de participação); 174,3 é reflexo da parcela AVJ implícita — a camada mais exposta. Os quadros RISCO 2 e RISCO 1 medem o IR adicional se o fisco negar uma ou as duas camadas (+59,3 e +149,2 em qualquer preço)."),
 ("CPC 19 (JO)","As holdings reconhecem 60% do VJ de todo o negócio (3.532,8): base contábil 472,6 + AVJ 3.060,2. O AVJ fica 'carimbado' a 34% (1.040,5) em subconta, com passivo diferido, e é devido na realização — piso da carga em qualquer preço. Acima do VJ, tributa-se 34% do excedente sobre a base contábil."),
 ("CPC 15 × CPC 19","No preço = VJ (3.533): 981 × 1.040 (PJ). A diferença é 34% × 174,3 — o CPC 19 equivale ao CPC 15 com a camada AVJ negada (RISCO 2), porque no CPC 15 o patrimônio da Embracon não é reavaliado e a mais-valia fica limitada ao PL da CNP."),
 ("Venda pela PF","No CPC 15, vender pelas PFs (22,5% sobre preço − 18) reduz a carga (791 × 981 no VJ), mas exige 'subir pra física' (redução de capital/liquidação das holdings — não modelado). No CPC 19, a PF não escapa do carimbo: paga 1.040 na holding mais 22,5% sobre o que exceder o VJ — por isso 'cai a tese' de vender pela PF."),
 ("Atualização 31.07 (draft2 v3)","Com PL de 254,0, VJ acordado de 5.954,8 e custo baixado de 66,6 (26,2%): custo 60% = 655,7 (187,5 + 293,9 + 174,3); JO contábil 481,3 e AVJ 3.091,6 (carimbo 1.051,1); venda ao VJ (3.572,9): CPC 15 991,9 × CPC 19 1.051,1 (PJ) e 799,8 × 1.051,1 (PF). Os degraus de risco passam a 59,3 (AVJ) e 99,9 (PL). Para reproduzir aqui, basta alterar as premissas azuis/vinculadas."),
]
r=68
for t,d in leit:
    put(f"B{r}",t,font=f_b,al=Alignment(horizontal="left",vertical="top",wrap_text=True))
    ws.merge_cells(f"C{r}:H{r}"); put(f"C{r}",d,al=Alignment(horizontal="left",vertical="top",wrap_text=True))
    for cc in "DEFGH": ws[f"{cc}{r}"].border=B
    ws.row_dimensions[r].height=44
    r+=1

wb.save(OUT); print("saved",OUT)
