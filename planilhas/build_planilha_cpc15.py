#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Planilha expositiva — Incorporação CNP x Embracon sob CPC 15.
Base de valores: "Exercicio_Incorporacao_draft1_MLA.xlsx", aba "base 30 06 26com dividendos".
Design: "Oportunidades_Intercompany_SF_26.08_V16.xlsx" (Calibri 9, barras C00000, zebra EAF0FB).
Todos os cálculos por fórmula, encadeados entre abas.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

OUT = "Exercicio_Incorporacao_CPC15_v1.xlsx"
LOGO = "xlsx_media_modelo/image1.png"

RED = "FFC00000"
ZEBRA = "FFEAF0FB"
NOTE = "FFF2F2F2"
GREEN = "FFE8F3DC"
AMBER = "FFFAF0DA"
GRAY_B = "FFC9C9C9"

f_title = Font(name="Calibri", size=12, bold=True)
f_bar = Font(name="Calibri", size=9, bold=True, color="FFFFFFFF")
f_bold = Font(name="Calibri", size=9, bold=True)
f_txt = Font(name="Calibri", size=9)
f_it = Font(name="Calibri", size=9, italic=True)
f_it_gray = Font(name="Calibri", size=9, italic=True, color="FF7F7F7F")
f_input = Font(name="Calibri", size=9, color="FF0000FF")
f_input_b = Font(name="Calibri", size=9, bold=True, color="FF0000FF")

fill_red = PatternFill("solid", fgColor=RED)
fill_zebra = PatternFill("solid", fgColor=ZEBRA)
fill_note = PatternFill("solid", fgColor=NOTE)
fill_green = PatternFill("solid", fgColor=GREEN)
fill_amber = PatternFill("solid", fgColor=AMBER)

thin = Side(style="thin", color=GRAY_B)
med = Side(style="medium", color=GRAY_B)
b_all = Border(left=thin, right=thin, top=thin, bottom=thin)

al_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
al_l = Alignment(horizontal="left", vertical="center", wrap_text=True)
al_lt = Alignment(horizontal="left", vertical="top", wrap_text=True)
al_r = Alignment(horizontal="right", vertical="center")

NF = "#,##0.00;[Red]\\(#,##0.00\\)"
NF1 = "#,##0.0;[Red]\\(#,##0.0\\)"
NF3 = "#,##0.000;[Red]\\(#,##0.000\\)"
PCT = "0.00%"
PCT1 = "0.0%"

wb = openpyxl.Workbook()

def put(ws, coord, value, font=f_txt, fill=None, nf=None, al=None, border=None):
    c = ws[coord]
    c.value = value
    c.font = font
    if fill: c.fill = fill
    if nf: c.number_format = nf
    c.alignment = al or (al_r if nf else al_l)
    if border: c.border = border
    return c

def bar(ws, rng, text):
    ws.merge_cells(rng)
    first = rng.split(":")[0]
    put(ws, first, text, font=f_bar, fill=fill_red, al=al_c)
    # borda em toda a faixa
    r1, r2 = rng.split(":")
    row = int("".join(ch for ch in r1 if ch.isdigit()))
    c1 = openpyxl.utils.column_index_from_string("".join(ch for ch in r1 if ch.isalpha()))
    c2 = openpyxl.utils.column_index_from_string("".join(ch for ch in r2 if ch.isalpha()))
    for cc in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=cc)
        cell.border = b_all
        cell.fill = fill_red

def box(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            left = med if c == c1 else thin
            right = med if c == c2 else thin
            top = med if r == r1 else thin
            bottom = med if r == r2 else thin
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

def borders(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = b_all

def note_box(ws, rng, text, fill=fill_note):
    ws.merge_cells(rng)
    first = rng.split(":")[0]
    put(ws, first, text, font=f_it, fill=fill, al=al_c)
    r1s, r2s = rng.split(":")
    row1 = int("".join(ch for ch in r1s if ch.isdigit()))
    row2 = int("".join(ch for ch in r2s if ch.isdigit()))
    c1 = openpyxl.utils.column_index_from_string("".join(ch for ch in r1s if ch.isalpha()))
    c2 = openpyxl.utils.column_index_from_string("".join(ch for ch in r2s if ch.isalpha()))
    for r in range(row1, row2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(row=r, column=cc).border = b_all
            ws.cell(row=r, column=cc).fill = fill

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

def comment(ws, coord, text):
    ws[coord].comment = Comment(text, "Análise CPC 15", width=320, height=140)

# =====================================================================
# ABA: Contexto
# =====================================================================
ws = wb.active
ws.title = "Contexto"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24, "G": 24, "H": 24})
try:
    img = Image(LOGO)
    img.width, img.height = 168, 40
    ws.add_image(img, "B2")
except Exception:
    pass
put(ws, "D2", "Projeto Another House — Exercício da Incorporação sob CPC 15", font=f_title)
put(ws, "D3", "Duas hipóteses de custo do investimento: custo original (\"magrinho\") x custo incrementado (\"gordinho\")", font=f_it)
put(ws, "B6", "Base 30.06.2026 (com dividendos) · valores em R$ milhões · draft de trabalho — não constitui opinião legal", font=f_it_gray)

bar(ws, "B8:H8", "O QUE ESTA PLANILHA DEMONSTRA")
note_box(ws, "B9:H11",
 "Demonstração, etapa por etapa, dos efeitos contábeis e fiscais da incorporação da CNP Consórcio pela Embracon sob a premissa de "
 "COMBINAÇÃO DE NEGÓCIOS (CPC 15), com atos societários lavrados a valor patrimonial contábil (laudo art. 224 LSA): (1) situação antes; "
 "(2) incorporação e reflexo nas holdings, com a segregação de risco do custo do investimento (MLA); (3) cash out; (4) destinação do caixa "
 "(PJ x PF); (5) alienação futura. Em cada etapa com efeito fiscal, comparam-se as duas hipóteses de custo.", fill_zebra)

bar(ws, "B13:H13", "AS DUAS HIPÓTESES DE CUSTO")
put(ws, "B14", "Hipótese A — custo original (\"magrinho\")", font=f_bold, fill=fill_amber, al=al_l, border=b_all)
ws.merge_cells("C14:H14")
put(ws, "C14", "O custo fiscal do investimento das holdings permanece o custo histórico (R$ 119,744 cada). A parcela incorporada (reflexo da "
 "incorporação) NÃO compõe custo — leitura de diferimento (jurisprudência recente do CARF: CPFL, Litela, Litel).", font=f_it, al=al_l, border=b_all)
for cc in "DEFGH": ws[f"{cc}14"].border = b_all
put(ws, "B15", "Hipótese B — custo incrementado (\"gordinho\")", font=f_bold, fill=fill_green, al=al_l, border=b_all)
ws.merge_cells("C15:H15")
put(ws, "C15", "A parcela incorporada líquida da diluição (R$ 344,970 por holding) integra o custo do investimento — leitura de exclusão "
 "definitiva do art. 33, §2º, do DL 1.598/77 (linha WTorre: o ganho neutralizado \"dá custo\").", font=f_it, al=al_l, border=b_all)
for cc in "DEFGH": ws[f"{cc}15"].border = b_all

bar(ws, "B17:H17", "COMO LER / LEGENDA")
put(ws, "B18", "Células em AZUL", font=f_input_b, border=b_all)
ws.merge_cells("C18:H18"); put(ws, "C18", "Entradas/premissas editáveis (hardcoded). Tudo o mais é fórmula encadeada.", font=f_txt, al=al_l, border=b_all)
for cc in "DEFGH": ws[f"{cc}18"].border = b_all
put(ws, "B19", "Células em PRETO", font=f_bold, border=b_all)
ws.merge_cells("C19:H19"); put(ws, "C19", "Fórmulas — o racional de cada número está na própria célula.", font=f_txt, al=al_l, border=b_all)
for cc in "DEFGH": ws[f"{cc}19"].border = b_all
put(ws, "B20", "Linhas \"Check\"", font=f_it_gray, border=b_all)
ws.merge_cells("C20:H20"); put(ws, "C20", "Conferências (devem ser ~zero). Diferenças herdadas da base do draft estão apontadas e não foram \"forçadas\".", font=f_txt, al=al_l, border=b_all)
for cc in "DEFGH": ws[f"{cc}20"].border = b_all

bar(ws, "B22:H22", "CRITÉRIOS E NOTAS RELEVANTES")
notas = [
 ("1.", "Valores adotados: aba \"base 30 06 26com dividendos\" do draft MLA (Exercicio_Incorporaçao_draft1). Balanços de 30.06.2026, em R$ milhões, já considerando dividendos de 120 declarados pela Embracon antes da incorporação."),
 ("2.", "Relação de troca: 13,8% para as Partes CNP (premissa do Acordo — EVs de R$ 821,9 mi x R$ 5.132,9 mi), e não a proporção dos PLs contábeis (22,2% x 77,8%). A incorporação é registrada a valor contábil (laudo art. 224 LSA)."),
 ("3.", "AJUSTE DE CRITÉRIO vs. draft: o custo da participação alienada no cash out foi rateado pela FRAÇÃO alienada do investimento (13,1 p.p. ÷ 43,1 p.p. = 30,39%). O draft multiplicava o custo por 13,1% diretamente (custo alocado de 15,69 vs 36,40 na Hipótese A) — critério ajustado por ser o rateio proporcional o tecnicamente defensável."),
 ("4.", "Cash out: preço de 600 por holding (1.200 total, sem correções do SPA). Tese da \"operação única\" (primária + secundária simultâneas): no cash out prevalece o custo antigo — a Hipótese A é a leitura conservadora nesta etapa."),
 ("5.", "Segregação de risco do custo (MLA): o novo custo das holdings é aberto em 3 camadas — (i) investimento antigo (perda de diluição); (ii) reflexo do PL contábil da CNP; (iii) reflexo da parcela AVJ implícita (proporção do exemplo Deloitte: MV 300 ÷ VJ 866 = 34,64%) — \"um tem mais risco que o outro\"."),
 ("6.", "Aba 4 (PJ x PF) usa as premissas do draft (IRRF dividendos 10%; CDI 12%; PJ a 100% do CDI com IR 34%; PF a 105% do CDI com IR 27%) — premissas a validar. Fórmula de rendimento corrigida para saldo × CDI × %CDI."),
 ("7.", "Checks de balanço: a base do draft tem diferenças de 0,34 (Embracon) e 0,84 (CNP) entre Ativos e Passivos+PL — mantidas e sinalizadas nas linhas \"Check\", a conciliar com o balancete."),
]
r = 23
for n, t in notas:
    put(ws, f"B{r}", n, font=f_bold, al=al_c, border=b_all)
    ws.merge_cells(f"C{r}:H{r}")
    put(ws, f"C{r}", t, font=f_txt, al=al_lt, border=b_all)
    for cc in "DEFGH": ws[f"{cc}{r}"].border = b_all
    if r % 2 == 1:
        for cc in "BCDEFGH": ws[f"{cc}{r}"].fill = fill_zebra
    ws.row_dimensions[r].height = 26
    r += 1

bar(ws, f"B{r+1}:H{r+1}", "FONTES")
note_box(ws, f"B{r+2}:H{r+3}",
 "Draft MLA (Exercicio_Incorporaçao_draft1, aba \"base 30 06 26com dividendos\") · Acordo de Investimento de 20.10.2025 (13,8%; R$ 1,2 bi; 60/40) · "
 "Call Deloitte 17.08.2026 (exemplo VJ 866 / MV 300) · Calls e e-mails de ago/2026 (hipóteses \"gordinho\" x \"magrinho\", segregação de risco, operação única).")
# alturas para textos longos
for rr in (9, 10, 11): ws.row_dimensions[rr].height = 22
ws.row_dimensions[14].height = 30
ws.row_dimensions[15].height = 30
for rr in range(23, 30): ws.row_dimensions[rr].height = 34
ws.row_dimensions[r + 2].height = 20
ws.row_dimensions[r + 3].height = 20

# =====================================================================
# ABA: Premissas
# =====================================================================
ws = wb.create_sheet("Premissas")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 52, "C": 14, "D": 64})
put(ws, "B1", "Premissas", font=f_title)
bar(ws, "B3:D3", "PREMISSAS GERAIS")
put(ws, "B4", "Data-base dos balanços", border=b_all); put(ws, "C4", "30.06.2026", font=f_input, al=al_c, border=b_all); put(ws, "D4", "Draft MLA — aba \"base 30 06 26com dividendos\"", font=f_it_gray, al=al_l, border=b_all)
put(ws, "B5", "Unidade", border=b_all); put(ws, "C5", "R$ milhões", font=f_input, al=al_c, border=b_all); put(ws, "D5", "", border=b_all)
put(ws, "B6", "Participação das Partes CNP na incorporação (relação de troca)", border=b_all); put(ws, "C6", 0.138, font=f_input, nf=PCT1, border=b_all); put(ws, "D6", "Acordo: EV CNP 821,9 ÷ (821,9 + 5.132,9) = 13,8% — definitiva e irrevisável", font=f_it_gray, al=al_l, border=b_all)
put(ws, "B7", "Participação-alvo final das Partes CNP", border=b_all); put(ws, "C7", 0.40, font=f_input, nf=PCT1, border=b_all); put(ws, "D7", "Acordo: 60% famílias (30/30) x 40% CNP", font=f_it_gray, al=al_l, border=b_all)
put(ws, "B8", "IRPJ/CSLL — ganho de capital PJ", border=b_all); put(ws, "C8", 0.34, font=f_input, nf=PCT1, border=b_all); put(ws, "D8", "", border=b_all)
put(ws, "B9", "Dividendos declarados pela Embracon antes da incorporação", border=b_all); put(ws, "C9", 120.0, font=f_input, nf=NF, border=b_all); put(ws, "D9", "Reduz o PL da Embracon a 239,488; permanece como passivo (a pagar às holdings)", font=f_it_gray, al=al_l, border=b_all)
for rr in range(4, 10):
    if rr % 2 == 0:
        for cc in "BCD": ws[f"{cc}{rr}"].fill = fill_zebra

bar(ws, "B11:D11", "SEGREGAÇÃO DO CUSTO — PROPORÇÃO DO EXEMPLO DELOITTE (ILUSTRATIVO)")
put(ws, "B12", "Valor justo do PL da CNP (exemplo DTT)", border=b_all); put(ws, "C12", 866.0, font=f_input, nf=NF, border=b_all); put(ws, "D12", "Call Deloitte 17.08.2026 — números ilustrativos, não contratuais", font=f_it_gray, al=al_l, border=b_all)
put(ws, "B13", "Mais-valia / AVJ contida no exemplo (MV)", border=b_all); put(ws, "C13", 300.0, font=f_input, nf=NF, border=b_all); put(ws, "D13", "Idem — parcela do VJ atribuída à reavaliação", font=f_it_gray, al=al_l, border=b_all)
put(ws, "B14", "% AVJ implícito no PL da CNP", font=f_bold, border=b_all); put(ws, "C14", "=C13/C12", font=f_bold, nf=PCT, border=b_all); put(ws, "D14", "MV ÷ VJ — aplicado ao PL contábil da CNP para segregar a camada de maior risco", font=f_it_gray, al=al_l, border=b_all)
put(ws, "B15", "% contábil (sem AVJ)", font=f_bold, border=b_all); put(ws, "C15", "=1-C14", font=f_bold, nf=PCT, border=b_all); put(ws, "D15", "", border=b_all)
for rr in (12, 14):
    for cc in "BCD": ws[f"{cc}{rr}"].fill = fill_zebra

bar(ws, "B17:D17", "CASH OUT")
put(ws, "B18", "Preço por holding (venda à CNP Participações)", border=b_all); put(ws, "C18", 600.0, font=f_input, nf=NF, border=b_all); put(ws, "D18", "Total 1.200 (R$ 1,2 bi do SPA, sem IPCA/ajustes de caixa) — 50/50 Savian/JVFJ", font=f_it_gray, al=al_l, border=b_all)
for cc in "BCD": ws["B18"].fill = fill_zebra; ws["C18"].fill = fill_zebra; ws["D18"].fill = fill_zebra

bar(ws, "B20:D20", "DESTINAÇÃO DO CAIXA E APLICAÇÕES (PREMISSAS DO DRAFT — A VALIDAR)")
rows20 = [
 ("IRRF sobre dividendos ao sócio PF", 0.10, "Premissa do draft (reforma tributária) — incide na distribuição da holding à PF"),
 ("CDI (a.a.)", 0.12, "Premissa do draft"),
 ("% do CDI — aplicação na PJ (holding)", 1.00, "Premissa do draft (100 do CDI)"),
 ("% do CDI — aplicação na PF", 1.05, "Premissa do draft (105 do CDI)"),
 ("IR sobre rendimentos — PJ", 0.34, "Premissa do draft (holding tributa o rendimento a 34%)"),
 ("IR sobre rendimentos — PF", 0.27, "Premissa do draft — a validar (renda fixa PF: 15%-22,5%)"),
 ("Horizonte de comparação (anos)", 2, "Premissa do draft"),
]
r = 21
for label, val, src in rows20:
    put(ws, f"B{r}", label, border=b_all)
    put(ws, f"C{r}", val, font=f_input, nf=(PCT1 if isinstance(val, float) else "0"), border=b_all)
    put(ws, f"D{r}", src, font=f_it_gray, al=al_l, border=b_all)
    if r % 2 == 1:
        for cc in "BCD": ws[f"{cc}{r}"].fill = fill_zebra
    r += 1
note_box(ws, f"B{r+1}:D{r+2}", "Todas as células azuis desta aba alimentam as demais por fórmula. Alterando uma premissa, a planilha inteira recalcula.")

P = "Premissas"  # atalho

# =====================================================================
# ABA: 1. Situação Antes
# =====================================================================
ws = wb.create_sheet("1. Situação Antes")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 26, "C": 12.5, "D": 2, "E": 26, "F": 12.5, "G": 2, "H": 26, "I": 12.5, "J": 2, "K": 26, "L": 12.5})
put(ws, "B1", "1. Situação antes da incorporação — balanços 30.06.2026", font=f_title)
bar(ws, "B2:L2", "BALANÇOS DE PARTIDA (R$ MILHÕES) — BASE DRAFT MLA \"COM DIVIDENDOS\"")

blocks = [("B", "C", "SAVIAN"), ("E", "F", "JVFJ"), ("H", "I", "EMBRACON"), ("K", "L", "CNP CONSÓRCIO")]
for cl, cv, name in blocks:
    ws.merge_cells(f"{cl}4:{cv}4")
    put(ws, f"{cl}4", name, font=f_bar, fill=fill_red, al=al_c)
    ws[f"{cv}4"].fill = fill_red
    for cc in (cl, cv): ws[f"{cc}4"].border = b_all

rows_lbl = {
 5: ("ATIVOS", True), 6: ("Circulante", False), 7: ("CPC 47 (despesas/comissões diferidas)", False),
 8: ("Permanente", False), 9: ("Inv. Embracon", False), 10: ("Compensação (grupos de consórcio)", False),
 12: ("PASSIVOS", True), 13: ("Circulante", False), 14: ("CPC 47 (receitas diferidas)", False),
 15: ("Dividendos a pagar", False), 16: ("Compensação (grupos de consórcio)", False),
 18: ("PATRIMÔNIO LÍQUIDO", True), 19: ("Capital", False), 20: ("Reservas", False), 21: ("(−) Dividendos declarados", False),
 23: ("PASSIVOS + PL", True), 24: ("Check (Ativos − Passivos − PL)", None),
}
for cl, cv, name in blocks:
    for rr, (lbl, bold) in rows_lbl.items():
        fnt = f_bold if bold else (f_it_gray if bold is None else f_txt)
        put(ws, f"{cl}{rr}", lbl, font=fnt, al=al_l)

# --- SAVIAN (C) e JVFJ (F): investimento = PL Embracon / 2
for cv in ("C", "F"):
    put(ws, f"{cv}5", f"=SUM({cv}6:{cv}10)", font=f_bold, nf=NF)
    put(ws, f"{cv}9", "=I18/2", nf=NF)
    comment(ws, f"{cv}9", "Custo/valor contábil do investimento na Embracon = 50% do PL da Embracon (premissa do draft: custo = valor patrimonial via MEP).")
    put(ws, f"{cv}12", f"=SUM({cv}13:{cv}16)", font=f_bold, nf=NF)
    put(ws, f"{cv}18", f"=SUM({cv}19:{cv}21)", font=f_bold, nf=NF)
    put(ws, f"{cv}19", f"={cv}9", nf=NF)
    put(ws, f"{cv}23", f"={cv}12+{cv}18", font=f_bold, nf=NF)
    put(ws, f"{cv}24", f"={cv}5-{cv}23", font=f_it_gray, nf=NF)

# --- EMBRACON (I)
put(ws, "I5", "=SUM(I6:I10)", font=f_bold, nf=NF)
put(ws, "I6", 328.578, font=f_input, nf=NF)
comment(ws, "I6", "Draft: 2.265,980 − 1.937,402 (reclassificação do ativo CPC 47 para linha própria).")
put(ws, "I7", 1937.402, font=f_input, nf=NF)
put(ws, "I8", 57.986, font=f_input, nf=NF)
put(ws, "I10", 23814.064, font=f_input, nf=NF)
put(ws, "I12", "=SUM(I13:I16)", font=f_bold, nf=NF)
put(ws, "I13", 420.060, font=f_input, nf=NF)
comment(ws, "I13", "Draft: 1.964,818 − 1.544,758 (reclassificação do passivo CPC 47 para linha própria).")
put(ws, "I14", 1544.758, font=f_input, nf=NF)
put(ws, "I15", f"={P}!C9", nf=NF)
put(ws, "I16", "=I10", nf=NF)
put(ws, "I18", "=SUM(I19:I21)", font=f_bold, nf=NF)
put(ws, "I19", 85.0, font=f_input, nf=NF)
put(ws, "I20", 274.488, font=f_input, nf=NF)
comment(ws, "I20", "Draft: 268,488 + 6,000.")
put(ws, "I21", f"=-{P}!C9", nf=NF)
put(ws, "I23", "=I12+I18", font=f_bold, nf=NF)
put(ws, "I24", "=I5-I23", font=f_it_gray, nf=NF)
comment(ws, "I24", "Diferença de −0,34 herdada da base do draft (Ativos 26.138,03 x Passivos+PL 26.138,37) — a conciliar com o balancete.")

# --- CNP (L)
put(ws, "L5", "=SUM(L6:L10)", font=f_bold, nf=NF)
put(ws, "L6", 945.234, font=f_input, nf=NF)
comment(ws, "L6", "Draft: 17.231,437 − 16.311,155 − 17,517 − 198,838 + 380 − 138,693 (Circulante + LP, líquido das reclassificações).")
put(ws, "L7", 216.353, font=f_input, nf=NF)
comment(ws, "L7", "Draft: 17,517 + 198,836.")
put(ws, "L8", 138.693, font=f_input, nf=NF)
put(ws, "L10", 16311.155, font=f_input, nf=NF)
put(ws, "L12", "=SUM(L13:L16)", font=f_bold, nf=NF)
put(ws, "L13", 232.757, font=f_input, nf=NF)
comment(ws, "L13", "Draft: 17.231,437 − 16.311 − 227,944 − 458,736 − 1.")
put(ws, "L14", 227.944, font=f_input, nf=NF)
put(ws, "L16", "=L10", nf=NF)
put(ws, "L18", "=SUM(L19:L21)", font=f_bold, nf=NF)
put(ws, "L19", 519.0, font=f_input, nf=NF)
comment(ws, "L19", "Draft: 139 + 380.")
put(ws, "L20", 319.735, font=f_input, nf=NF)
comment(ws, "L20", "Draft: 251,958 + 67,777.")
put(ws, "L23", "=L12+L18", font=f_bold, nf=NF)
put(ws, "L24", "=L5-L23", font=f_it_gray, nf=NF)
comment(ws, "L24", "Diferença de +0,84 herdada da base do draft — a conciliar com o balancete.")

for cl, cv, name in blocks:
    c1 = openpyxl.utils.column_index_from_string(cl)
    box(ws, 4, c1, 24, c1 + 1)
    for rr in (5, 12, 18, 23):
        for cc in (cl, cv): ws[f"{cc}{rr}"].fill = fill_zebra

bar(ws, "B27:F27", "PROPORCIONALIDADE DOS PLs \"NOMINAIS\"")
put(ws, "B28", "Entidade", font=f_bold, fill=fill_zebra, border=b_all); put(ws, "C28", "PL 30.06.26", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
ws.merge_cells("D28:E28"); put(ws, "D28", "% do PL combinado", font=f_bold, fill=fill_zebra, al=al_c, border=b_all); ws["E28"].border = b_all; ws["E28"].fill = fill_zebra
put(ws, "F28", "", border=b_all); ws["F28"].fill = fill_zebra
put(ws, "B29", "Embracon", border=b_all); put(ws, "C29", "=I18", nf=NF, border=b_all)
ws.merge_cells("D29:E29"); put(ws, "D29", "=C29/C31", nf=PCT, al=al_c, border=b_all); ws["E29"].border = b_all
put(ws, "F29", "", border=b_all)
put(ws, "B30", "CNP Consórcio", border=b_all); put(ws, "C30", "=L18", nf=NF, border=b_all)
ws.merge_cells("D30:E30"); put(ws, "D30", "=C30/C31", nf=PCT, al=al_c, border=b_all); ws["E30"].border = b_all
put(ws, "F30", "", border=b_all)
put(ws, "B31", "PL combinado", font=f_bold, fill=fill_zebra, border=b_all); put(ws, "C31", "=C29+C30", font=f_bold, nf=NF, fill=fill_zebra, border=b_all)
ws.merge_cells("D31:E31"); put(ws, "D31", "=C31/C31", font=f_bold, nf=PCT, al=al_c, fill=fill_zebra, border=b_all); ws["E31"].border = b_all; ws["E31"].fill = fill_zebra
put(ws, "F31", "", border=b_all); ws["F31"].fill = fill_zebra
note_box(ws, "B33:L34",
 "A relação de troca contratual (13,8% / 86,2%) NÃO segue a proporção dos PLs contábeis acima (22,2% / 77,8%): decorre dos EVs acordados no "
 "Acordo (CNP 821,9 x Embracon 5.132,9). É exatamente esse descolamento que gera a \"parcela incorporada\" nas holdings e a discussão de custo.")

SA = "'1. Situação Antes'"

# =====================================================================
# ABA: 2. Incorporação CPC 15
# =====================================================================
ws = wb.create_sheet("2. Incorporação CPC 15")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 40, "C": 14, "D": 14, "E": 14, "F": 14, "G": 2, "H": 24, "I": 13, "J": 13, "K": 13, "L": 13})
put(ws, "B1", "2. Incorporação da CNP pela Embracon (CPC 15 — atos a valor contábil)", font=f_title)

bar(ws, "B3:F3", "RELAÇÃO DE TROCA E PARTICIPAÇÕES")
put(ws, "B4", "Acionista", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C4", "Antes", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "D4", "Após incorporação", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "E4", "", border=b_all); ws["E4"].fill = fill_zebra
put(ws, "F4", "", border=b_all); ws["F4"].fill = fill_zebra
put(ws, "B5", "Cia. Savian", border=b_all); put(ws, "C5", 0.50, font=f_input, nf=PCT1, border=b_all)
put(ws, "D5", f"=C5*(1-{P}!C6)", nf=PCT, border=b_all); put(ws, "E5", "", border=b_all); put(ws, "F5", "", border=b_all)
put(ws, "B6", "Cia. JVFJ", border=b_all); put(ws, "C6", 0.50, font=f_input, nf=PCT1, border=b_all)
put(ws, "D6", f"=C6*(1-{P}!C6)", nf=PCT, border=b_all); put(ws, "E6", "", border=b_all); put(ws, "F6", "", border=b_all)
put(ws, "B7", "Partes CNP (ações novas)", border=b_all); put(ws, "C7", 0.0, font=f_input, nf=PCT1, border=b_all)
put(ws, "D7", f"={P}!C6", nf=PCT, border=b_all); put(ws, "E7", "", border=b_all); put(ws, "F7", "", border=b_all)
put(ws, "B8", "Total", font=f_bold, fill=fill_zebra, border=b_all); put(ws, "C8", "=SUM(C5:C7)", font=f_bold, nf=PCT1, fill=fill_zebra, border=b_all)
put(ws, "D8", "=SUM(D5:D7)", font=f_bold, nf=PCT1, fill=fill_zebra, border=b_all)
put(ws, "E8", "", border=b_all); ws["E8"].fill = fill_zebra
put(ws, "F8", "", border=b_all); ws["F8"].fill = fill_zebra

bar(ws, "H3:L3", "EMBRACON COMBINADA — BALANÇO APÓS (VALOR CONTÁBIL)")
put(ws, "H4", "Linha", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "I4", "Embracon", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "J4", "CNP", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "K4", "Combinada", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "L4", "", border=b_all); ws["L4"].fill = fill_zebra
linhas_comb = [
 ("Ativo circulante", "I6", "L6"), ("Ativo CPC 47", "I7", "L7"), ("Permanente", "I8", "L8"), ("Compensação (ativo)", "I10", "L10"),
 ("Passivo circulante", "I13", "L13"), ("Passivo CPC 47", "I14", "L14"), ("Dividendos a pagar", "I15", "L15"), ("Compensação (passivo)", "I16", "L16"),
 ("Capital social", "I19", "L19"), ("Reservas", "I20", "L20"), ("(−) Dividendos", "I21", "L21"),
]
r = 5
for lbl, ce, cc_ in linhas_comb:
    put(ws, f"H{r}", lbl, border=b_all)
    put(ws, f"I{r}", f"={SA}!{ce}", nf=NF, border=b_all)
    put(ws, f"J{r}", f"={SA}!{cc_}", nf=NF, border=b_all)
    put(ws, f"K{r}", f"=I{r}+J{r}", nf=NF, border=b_all)
    put(ws, f"L{r}", "", border=b_all)
    r += 1
put(ws, f"H{r}", "PL combinado", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, f"I{r}", f"={SA}!I18", font=f_bold, nf=NF, fill=fill_zebra, border=b_all)
put(ws, f"J{r}", f"={SA}!L18", font=f_bold, nf=NF, fill=fill_zebra, border=b_all)
put(ws, f"K{r}", f"=I{r}+J{r}", font=f_bold, nf=NF, fill=fill_zebra, border=b_all)
put(ws, f"L{r}", "", border=b_all); ws[f"L{r}"].fill = fill_zebra
PLCOMB = f"K{r}"  # K16
r += 1
put(ws, f"H{r}", "Check (A − P − PL)", font=f_it_gray, border=b_all)
put(ws, f"I{r}", "=SUM(I5:I8)-SUM(I9:I12)-I16", font=f_it_gray, nf=NF, border=b_all)
put(ws, f"J{r}", "=SUM(J5:J8)-SUM(J9:J12)-J16", font=f_it_gray, nf=NF, border=b_all)
put(ws, f"K{r}", "=SUM(K5:K8)-SUM(K9:K12)-K16", font=f_it_gray, nf=NF, border=b_all)
put(ws, f"L{r}", "", border=b_all)

bar(ws, "B10:F10", "REFLEXO NAS HOLDINGS (MEP) — CUSTO DO INVESTIMENTO")
put(ws, "B11", "Item", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C11", "Por holding", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "D11", "Total (2 holdings)", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "E11", "", border=b_all); ws["E11"].fill = fill_zebra
put(ws, "F11", "", border=b_all); ws["F11"].fill = fill_zebra
put(ws, "B12", "Investimento anterior (custo original — \"magrinho\")", border=b_all)
put(ws, "C12", f"={SA}!C9", nf=NF, border=b_all); put(ws, "D12", "=C12*2", nf=NF, border=b_all)
put(ws, "E12", "", border=b_all); put(ws, "F12", "", border=b_all)
put(ws, "B13", "(+) Incorporação base (reflexo do PL contábil da CNP ÷ 2)", border=b_all)
put(ws, "C13", f"={SA}!L18/2", nf=NF, border=b_all); put(ws, "D13", "=C13*2", nf=NF, border=b_all)
put(ws, "E13", "", border=b_all); put(ws, "F13", "", border=b_all)
put(ws, "B14", "(−) Diluição (13,8% sobre investimento + base)", border=b_all)
put(ws, "C14", f"=-(C12+C13)*{P}!C6", nf=NF, border=b_all); put(ws, "D14", "=C14*2", nf=NF, border=b_all)
put(ws, "E14", "", border=b_all); put(ws, "F14", "", border=b_all)
put(ws, "B15", "(=) Parcela incorporada (líquida)", font=f_bold, border=b_all)
put(ws, "C15", "=C13+C14", font=f_bold, nf=NF, border=b_all); put(ws, "D15", "=C15*2", font=f_bold, nf=NF, border=b_all)
put(ws, "E15", "", border=b_all); put(ws, "F15", "", border=b_all)
put(ws, "B16", "(=) Custo final do investimento (\"gordinho\")", font=f_bold, fill=fill_green, border=b_all)
put(ws, "C16", "=C12+C15", font=f_bold, nf=NF, fill=fill_green, border=b_all)
put(ws, "D16", "=C16*2", font=f_bold, nf=NF, fill=fill_green, border=b_all)
put(ws, "E16", "", border=b_all); ws["E16"].fill = fill_green
put(ws, "F16", "", border=b_all); ws["F16"].fill = fill_green
put(ws, "B17", "Check: 43,1% × PL combinado", font=f_it_gray, border=b_all)
put(ws, "C17", f"=D5*{PLCOMB}-C16", font=f_it_gray, nf=NF, border=b_all)
put(ws, "D17", "", border=b_all); put(ws, "E17", "", border=b_all); put(ws, "F17", "", border=b_all)

bar(ws, "B19:F19", "LANÇAMENTOS CONTÁBEIS NAS HOLDINGS (CADA UMA)")
put(ws, "B20", "Lançamento", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C20", "Débito", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "D20", "Crédito", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
ws.merge_cells("E20:F20"); put(ws, "E20", "Contrapartida / natureza", font=f_bold, fill=fill_zebra, al=al_c, border=b_all); ws["F20"].border = b_all; ws["F20"].fill = fill_zebra
put(ws, "B21", "1. Reflexo da incorporação — Inv. Embracon", border=b_all)
put(ws, "C21", "=C13", nf=NF, border=b_all); put(ws, "D21", "", border=b_all)
ws.merge_cells("E21:F21"); put(ws, "E21", "Conta patrimonial (ORA x reserva — EM ABERTO; CPC não define; Deloitte não respondeu)", font=f_it, al=al_l, border=b_all); ws["F21"].border = b_all
put(ws, "B22", "2. Perda de diluição — Inv. Embracon", border=b_all)
put(ws, "C22", "", border=b_all); put(ws, "D22", "=-C14", nf=NF, border=b_all)
ws.merge_cells("E22:F22"); put(ws, "E22", "Resultado de equivalência (variação de participação, art. 33, §2º, DL 1.598/77)", font=f_it, al=al_l, border=b_all); ws["F22"].border = b_all
put(ws, "B23", "Efeito líquido no investimento", font=f_bold, border=b_all)
put(ws, "C23", "=C21-D22", font=f_bold, nf=NF, border=b_all); put(ws, "D23", "", border=b_all)
ws.merge_cells("E23:F23"); put(ws, "E23", "(igual à parcela incorporada)", font=f_it_gray, al=al_l, border=b_all); ws["F23"].border = b_all

bar(ws, "B25:F25", "SEGREGAÇÃO DE RISCO DO CUSTO (MLA) — 3 CAMADAS")
put(ws, "B26", "Camada", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C26", "PL combinado", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "D26", "Antes (86,2%→)", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "E26", "Depois (×86,2%)", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "F26", "Δ (2 holdings)", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "B27", "1. Investimento antigo (PL Embracon) — perda de diluição", border=b_all)
put(ws, "C27", f"={SA}!I18", nf=NF, border=b_all)
put(ws, "D27", "=C27", nf=NF, border=b_all)
put(ws, "E27", f"=C27*(1-{P}!C6)", nf=NF, border=b_all)
put(ws, "F27", "=E27-D27", nf=NF, border=b_all)
put(ws, "B28", "2. Reflexo do PL contábil da CNP (menor risco)", border=b_all)
put(ws, "C28", f"={SA}!L18*{P}!C15", nf=NF, border=b_all)
put(ws, "D28", 0.0, nf=NF, border=b_all)
put(ws, "E28", f"=C28*(1-{P}!C6)", nf=NF, border=b_all)
put(ws, "F28", "=E28-D28", nf=NF, border=b_all)
put(ws, "B29", "3. Reflexo da parcela AVJ implícita (maior risco)", fill=fill_amber, border=b_all)
put(ws, "C29", f"={SA}!L18*{P}!C14", nf=NF, fill=fill_amber, border=b_all)
put(ws, "D29", 0.0, nf=NF, fill=fill_amber, border=b_all)
put(ws, "E29", f"=C29*(1-{P}!C6)", nf=NF, fill=fill_amber, border=b_all)
put(ws, "F29", "=E29-D29", nf=NF, fill=fill_amber, border=b_all)
put(ws, "B30", "Novo Embracon (soma)", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C30", "=SUM(C27:C29)", font=f_bold, nf=NF, fill=fill_zebra, border=b_all)
put(ws, "D30", "=SUM(D27:D29)", font=f_bold, nf=NF, fill=fill_zebra, border=b_all)
put(ws, "E30", "=SUM(E27:E29)", font=f_bold, nf=NF, fill=fill_zebra, border=b_all)
put(ws, "F30", "=SUM(F27:F29)", font=f_bold, nf=NF, fill=fill_zebra, border=b_all)
put(ws, "B31", "Δ por holding (÷2) — deve igualar a parcela incorporada", font=f_it_gray, border=b_all)
put(ws, "C31", "", border=b_all); put(ws, "D31", "", border=b_all)
put(ws, "E31", "", border=b_all)
put(ws, "F31", "=F30/2-C15", font=f_it_gray, nf=NF, border=b_all)
comment(ws, "B29", "Proporção AVJ = MV 300 ÷ VJ 866 = 34,64% (exemplo Deloitte, call 17.08.2026) aplicada ao PL contábil da CNP. Camada com maior risco de questionamento (AVJ Reflexo).")

bar(ws, "H19:L19", "LADO CNP HOLDING (INFORMATIVO)")
put(ws, "H20", "Baixa do acervo vertido (PL CNP)", border=b_all)
put(ws, "I20", f"=-{SA}!L18", nf=NF, border=b_all)
for cc in "JKL": put(ws, f"{cc}20", "", border=b_all)
put(ws, "H21", "Participação recebida (13,8% × PL combinado)", border=b_all)
put(ws, "I21", f"={P}!C6*{PLCOMB}", nf=NF, border=b_all)
for cc in "JKL": put(ws, f"{cc}21", "", border=b_all)
put(ws, "H22", "Resultado contábil a valor contábil", font=f_bold, border=b_all)
put(ws, "I22", "=I20+I21", font=f_bold, nf=NF, border=b_all)
for cc in "JKL": put(ws, f"{cc}22", "", border=b_all)
ws.merge_cells("H23:L24")
put(ws, "H23", "Espelho da parcela incorporada das holdings (−689,94). É esta perda a valor contábil que leva a CNP a defender a joint operation "
 "(CPC 19) com remensuração a valor justo em resultado.", font=f_it, fill=fill_note, al=al_c)
for rr in range(23, 25):
    for cc in "HIJKL": ws[f"{cc}{rr}"].border = b_all; ws[f"{cc}{rr}"].fill = fill_note

note_box(ws, "B33:L34",
 "Sob CPC 15 \"pleno\" haveria PPA com laudo a valor justo do PL da CNP (mais-valia \"grudando\" nos ativos). Neste exercício, seguindo o draft e o "
 "contrato (laudo a valor patrimonial contábil, art. 224 LSA), a incorporação é registrada a valor contábil e a dimensão CPC 15 aparece na segregação "
 "de risco acima — premissa da Malu: todo o excedente é mais-valia/AVJ, sem goodwill.")
ws.row_dimensions[33].height = 20
ws.row_dimensions[34].height = 20
ws.row_dimensions[23].height = 20
ws.row_dimensions[24].height = 20

INC = "'2. Incorporação CPC 15'"

# =====================================================================
# ABA: 3. Cash Out
# =====================================================================
ws = wb.create_sheet("3. Cash Out")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 44, "C": 15, "D": 15, "E": 2, "F": 15, "G": 15})
put(ws, "B1", "3. Cash out — venda de 26,2% à CNP (etapa 2, simultânea)", font=f_title)

bar(ws, "B3:G3", "PARÂMETROS DA ETAPA")
put(ws, "B4", "Pontos percentuais vendidos — total (40% − 13,8%)", border=b_all)
put(ws, "C4", f"={P}!C7-{P}!C6", nf=PCT, border=b_all)
for cc in "DFG": put(ws, f"{cc}4", "", border=b_all)
put(ws, "E4", "", border=b_all)
put(ws, "B5", "Pontos percentuais vendidos — por holding", border=b_all)
put(ws, "C5", "=C4/2", nf=PCT, border=b_all)
for cc in "DFG": put(ws, f"{cc}5", "", border=b_all)
put(ws, "E5", "", border=b_all)
put(ws, "B6", "Participação de cada holding após a incorporação", border=b_all)
put(ws, "C6", f"={INC}!D5", nf=PCT, border=b_all)
for cc in "DFG": put(ws, f"{cc}6", "", border=b_all)
put(ws, "E6", "", border=b_all)
put(ws, "B7", "Fração alienada do investimento (13,1 ÷ 43,1)", font=f_bold, border=b_all)
put(ws, "C7", "=C5/C6", font=f_bold, nf=PCT, border=b_all)
for cc in "DFG": put(ws, f"{cc}7", "", border=b_all)
put(ws, "E7", "", border=b_all)
comment(ws, "C7", "Critério ajustado vs. draft: rateio do custo pela fração alienada do investimento (13,1/43,1 = 30,39%), e não 13,1% direto. Ver nota 3 do Contexto.")
put(ws, "B8", "Preço por holding", border=b_all)
put(ws, "C8", f"={P}!C18", nf=NF, border=b_all)
for cc in "DFG": put(ws, f"{cc}8", "", border=b_all)
put(ws, "E8", "", border=b_all)

bar(ws, "B10:G10", "PARTICIPAÇÕES: APÓS INCORPORAÇÃO → VENDA → FINAL")
put(ws, "B11", "Acionista", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C11", "Após incorp.", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "D11", "Venda", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "E11", "", border=b_all); ws["E11"].fill = fill_zebra
put(ws, "F11", "Final", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "G11", "", border=b_all); ws["G11"].fill = fill_zebra
part_rows = [("Cia. Savian", f"={INC}!D5", "=-$C$5"), ("Cia. JVFJ", f"={INC}!D6", "=-$C$5"), ("Partes CNP", f"={INC}!D7", "=$C$4")]
r = 12
for nome, apos, venda in part_rows:
    put(ws, f"B{r}", nome, border=b_all)
    put(ws, f"C{r}", apos, nf=PCT, border=b_all)
    put(ws, f"D{r}", venda, nf=PCT, border=b_all)
    put(ws, f"E{r}", "", border=b_all)
    put(ws, f"F{r}", f"=C{r}+D{r}", nf=PCT, border=b_all)
    put(ws, f"G{r}", "", border=b_all)
    r += 1
put(ws, f"B{r}", "Total", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, f"C{r}", "=SUM(C12:C14)", font=f_bold, nf=PCT1, fill=fill_zebra, border=b_all)
put(ws, f"D{r}", "=SUM(D12:D14)", font=f_bold, nf=PCT1, fill=fill_zebra, border=b_all)
put(ws, f"E{r}", "", border=b_all); ws[f"E{r}"].fill = fill_zebra
put(ws, f"F{r}", "=SUM(F12:F14)", font=f_bold, nf=PCT1, fill=fill_zebra, border=b_all)
put(ws, f"G{r}", "", border=b_all); ws[f"G{r}"].fill = fill_zebra

bar(ws, "B17:G17", "GANHO DE CAPITAL — DUAS HIPÓTESES DE CUSTO")
ws.merge_cells("C18:D18"); put(ws, "C18", "Por holding", font=f_bold, fill=fill_zebra, al=al_c, border=b_all); ws["D18"].border = b_all; ws["D18"].fill = fill_zebra
put(ws, "E18", "", border=b_all)
ws.merge_cells("F18:G18"); put(ws, "F18", "Total (2 holdings)", font=f_bold, fill=fill_zebra, al=al_c, border=b_all); ws["G18"].border = b_all; ws["G18"].fill = fill_zebra
put(ws, "B18", "", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "B19", "Item", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C19", "Hip. A — custo original", font=f_bold, fill=fill_amber, al=al_r, border=b_all)
put(ws, "D19", "Hip. B — custo incrementado", font=f_bold, fill=fill_green, al=al_r, border=b_all)
put(ws, "E19", "", border=b_all)
put(ws, "F19", "Hip. A", font=f_bold, fill=fill_amber, al=al_r, border=b_all)
put(ws, "G19", "Hip. B", font=f_bold, fill=fill_green, al=al_r, border=b_all)
gk_rows = [
 ("Preço da alienação", "=$C$8", "=$C$8"),
 ("Custo total do investimento", f"={INC}!C12", f"={INC}!C16"),
 ("Custo alocado à parcela alienada (× fração 30,39%)", "=C21*$C$7", "=D21*$C$7"),
 ("Ganho de capital", "=C20-C22", "=D20-D22"),
 ("IRPJ/CSLL (34%)", f"=C23*{P}!C8", f"=D23*{P}!C8"),
 ("Líquido do imposto (preço − IR)", "=C20-C24", "=D20-D24"),
 ("Alíquota efetiva sobre o preço", "=C24/C20", "=D24/D20"),
 ("Custo remanescente (para a alienação futura)", "=C21-C22", "=D21-D22"),
]
r = 20
for lbl, fa, fb in gk_rows:
    bold = lbl.startswith(("Ganho", "IRPJ", "Custo remanescente"))
    fnt = f_bold if bold else f_txt
    nf_use = PCT if "Alíquota" in lbl else NF
    put(ws, f"B{r}", lbl, font=fnt, al=al_l, border=b_all)
    put(ws, f"C{r}", fa, font=fnt, nf=nf_use, border=b_all)
    put(ws, f"D{r}", fb, font=fnt, nf=nf_use, border=b_all)
    put(ws, f"E{r}", "", border=b_all)
    if "Alíquota" in lbl:
        put(ws, f"F{r}", "=F24/F20", font=fnt, nf=PCT, border=b_all)
        put(ws, f"G{r}", "=G24/G20", font=fnt, nf=PCT, border=b_all)
    else:
        put(ws, f"F{r}", f"=C{r}*2", font=fnt, nf=nf_use, border=b_all)
        put(ws, f"G{r}", f"=D{r}*2", font=fnt, nf=nf_use, border=b_all)
    r += 1
put(ws, f"B{r}", "Economia de IR da Hipótese B nesta etapa", font=f_bold, fill=fill_green, border=b_all)
put(ws, f"C{r}", "=C24-D24", font=f_bold, nf=NF, fill=fill_green, border=b_all)
put(ws, f"D{r}", "", border=b_all); ws[f"D{r}"].fill = fill_green
put(ws, f"E{r}", "", border=b_all)
put(ws, f"F{r}", "=F24-G24", font=f_bold, nf=NF, fill=fill_green, border=b_all)
put(ws, f"G{r}", "", border=b_all); ws[f"G{r}"].fill = fill_green

note_box(ws, f"B{r+2}:G{r+4}",
 "Tese da \"operação única\": incorporação e cash out simultâneos formam um só negócio — no cash out prevaleceria o CUSTO ANTIGO (Hipótese A), "
 "sem aproveitamento da parcela incorporada. A Hipótese B pressupõe que a parcela incorporada compõe custo já nesta etapa (art. 33, §2º, como exclusão "
 "definitiva — WTorre). Preço de 600 por holding sem correções do SPA (IPCA + ajustes de caixa líquido + earn-out de até 680 não modelados).")

CO = "'3. Cash Out'"

# =====================================================================
# ABA: 4. Caixa PJ x PF
# =====================================================================
ws = wb.create_sheet("4. Caixa PJ x PF")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 46, "C": 16, "D": 16, "E": 40})
put(ws, "B1", "4. Destinação do caixa do cash out — manter na PJ x distribuir à PF", font=f_title)

bar(ws, "B3:E3", "PONTO DE PARTIDA (POR HOLDING — HIPÓTESE A, CONSERVADORA)")
put(ws, "B4", "Líquido do cash out na holding (após IR de 34%)", border=b_all)
put(ws, "C4", f"={CO}!C25", nf=NF, border=b_all)
put(ws, "D4", "", border=b_all)
put(ws, "E4", "Aba 3, Hipótese A (custo original)", font=f_it_gray, al=al_l, border=b_all)
put(ws, "B5", "IRRF na distribuição de dividendos à PF (10%)", border=b_all)
put(ws, "C5", f"=C4*{P}!C21", nf=NF, border=b_all)
put(ws, "D5", "", border=b_all)
put(ws, "E5", "Premissa do draft (reforma tributária)", font=f_it_gray, al=al_l, border=b_all)
put(ws, "B6", "Líquido disponível na PF", font=f_bold, border=b_all)
put(ws, "C6", "=C4-C5", font=f_bold, nf=NF, border=b_all)
put(ws, "D6", "", border=b_all)
put(ws, "E6", "", border=b_all)

bar(ws, "B8:E8", "APLICAÇÃO FINANCEIRA — 2 ANOS (PREMISSAS DO DRAFT)")
put(ws, "B9", "Item", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C9", "Manter na PJ", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "D9", "Distribuir à PF", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
put(ws, "E9", "Racional", font=f_bold, fill=fill_zebra, al=al_l, border=b_all)
lin = [
 ("Saldo inicial", "=C4", "=C6", "PJ mantém 100%; PF recebe líquido do IRRF de 10%"),
 ("Ano 1 — rendimento bruto", f"=C10*{P}!C22*{P}!C23", f"=D10*{P}!C22*{P}!C24", "saldo × CDI × % do CDI (PJ 100% · PF 105%)"),
 ("Ano 1 — IR sobre o rendimento", f"=-C11*{P}!C25", f"=-D11*{P}!C26", "PJ 34% · PF 27% (draft — a validar)"),
 ("Saldo ao fim do ano 1", "=SUM(C10:C12)", "=SUM(D10:D12)", ""),
 ("Ano 2 — rendimento bruto", f"=C13*{P}!C22*{P}!C23", f"=D13*{P}!C22*{P}!C24", ""),
 ("Ano 2 — IR sobre o rendimento", f"=-C14*{P}!C25", f"=-D14*{P}!C26", ""),
 ("Saldo ao fim do ano 2", "=SUM(C13:C15)", "=SUM(D13:D15)", ""),
]
r = 10
for lbl, fc, fd, rac in lin:
    bold = lbl.startswith("Saldo")
    fnt = f_bold if bold else f_txt
    put(ws, f"B{r}", lbl, font=fnt, al=al_l, border=b_all)
    put(ws, f"C{r}", fc, font=fnt, nf=NF, border=b_all)
    put(ws, f"D{r}", fd, font=fnt, nf=NF, border=b_all)
    put(ws, f"E{r}", rac, font=f_it_gray, al=al_l, border=b_all)
    if bold:
        for cc in "BCDE": ws[f"{cc}{r}"].fill = fill_zebra
    r += 1
put(ws, f"B{r}", "Vantagem de manter na PJ (fim do ano 2)", font=f_bold, fill=fill_green, border=b_all)
put(ws, f"C{r}", "=C16-D16", font=f_bold, nf=NF, fill=fill_green, border=b_all)
put(ws, f"D{r}", "", border=b_all); ws[f"D{r}"].fill = fill_green
put(ws, f"E{r}", "", border=b_all); ws[f"E{r}"].fill = fill_green

note_box(ws, f"B{r+2}:E{r+4}",
 "Leitura: distribuir à PF custa 10% de IRRF na largada; a PF aplica a 105% do CDI com IR menor (27% vs 34%), mas em 2 anos não recupera a perda "
 "inicial — a PJ termina à frente. Premissas do draft a validar (IR de PF em renda fixa é 15%–22,5% conforme prazo; a comparação é sensível a essas "
 "alíquotas e ao horizonte). Sob CPC 19/AVJ diferido, \"subir pra física\" fica ainda mais custoso (ganho carimbado nas holdings).")

# =====================================================================
# ABA: 5. Alienação Futura
# =====================================================================
ws = wb.create_sheet("5. Alienação Futura")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 44, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13})
put(ws, "B1", "5. Alienação futura (\"passo F\") — sensibilidade de preço", font=f_title)

bar(ws, "B3:G3", "BASE (POR HOLDING)")
put(ws, "B4", "Participação remanescente", border=b_all)
put(ws, "C4", f"={CO}!F12", nf=PCT, border=b_all)
for cc in "DEFG": put(ws, f"{cc}4", "", border=b_all)
put(ws, "B5", "Custo remanescente — Hip. A (custo original)", border=b_all)
put(ws, "C5", f"={CO}!C27", nf=NF, border=b_all)
for cc in "DEFG": put(ws, f"{cc}5", "", border=b_all)
put(ws, "B6", "Custo remanescente — Hip. B (custo incrementado)", border=b_all)
put(ws, "C6", f"={CO}!D27", nf=NF, border=b_all)
for cc in "DEFG": put(ws, f"{cc}6", "", border=b_all)
put(ws, "B7", "Diferença de custo (B − A)", font=f_bold, border=b_all)
put(ws, "C7", "=C6-C5", font=f_bold, nf=NF, border=b_all)
for cc in "DEFG": put(ws, f"{cc}7", "", border=b_all)

bar(ws, "B9:G9", "SENSIBILIDADE — PREÇO DA VENDA FUTURA (POR HOLDING, R$ MILHÕES)")
put(ws, "B10", "Preço da venda futura (editável)", font=f_input_b, fill=fill_zebra, border=b_all)
precos = [300, 600, 900, 1200, 1500]
for i, p_ in enumerate(precos):
    cl = get_column_letter(3 + i)
    put(ws, f"{cl}10", float(p_), font=f_input_b, nf=NF1, fill=fill_zebra, border=b_all)
sens_rows = [
 ("Ganho de capital — Hip. A", "={c}10-$C$5", NF, False, None),
 ("IRPJ/CSLL 34% — Hip. A", f"=MAX(0,{{c}}11)*{P}!C8", NF, True, fill_amber),
 ("Ganho de capital — Hip. B", "={c}10-$C$6", NF, False, None),
 ("IRPJ/CSLL 34% — Hip. B", f"=MAX(0,{{c}}13)*{P}!C8", NF, True, fill_green),
 ("Economia de IR da Hip. B (A − B)", "={c}12-{c}14", NF, True, fill_green),
 ("Idem, total 2 holdings", "={c}15*2", NF, True, None),
]
r = 11
for lbl, tmpl, nf_use, bold, fl in sens_rows:
    fnt = f_bold if bold else f_txt
    put(ws, f"B{r}", lbl, font=fnt, al=al_l, fill=fl, border=b_all)
    for i in range(len(precos)):
        cl = get_column_letter(3 + i)
        put(ws, f"{cl}{r}", tmpl.format(c=cl), font=fnt, nf=nf_use, fill=fl, border=b_all)
    r += 1
note_box(ws, f"B{r+1}:G{r+3}",
 "A economia da Hipótese B tende ao teto de 34% × diferença de custo (81,64 por holding; 163,28 nas duas) e independe do preço enquanto o ganho for "
 "positivo nas duas hipóteses (preços baixos limitam o efeito — IR não fica negativo). Não há preço definido para o passo F: a linha de preço é "
 "editável. Sob CPC 19, o AVJ fica \"carimbado\" a 34% em subconta — economicamente equivalente à Hipótese A, com tributação garantida na realização.")

AF = "'5. Alienação Futura'"

# =====================================================================
# ABA: Conclusão
# =====================================================================
ws = wb.create_sheet("Conclusão")
ws.sheet_view.showGridLines = False
widths(ws, {"A": 2.5, "B": 52, "C": 16, "D": 16, "E": 16})
put(ws, "B1", "Conclusão — síntese das duas hipóteses de custo", font=f_title)

bar(ws, "B3:E3", "SÍNTESE (TOTAL DAS 2 HOLDINGS, R$ MILHÕES)")
put(ws, "B4", "Item", font=f_bold, fill=fill_zebra, border=b_all)
put(ws, "C4", "Hip. A — custo original", font=f_bold, fill=fill_amber, al=al_r, border=b_all)
put(ws, "D4", "Hip. B — custo incrementado", font=f_bold, fill=fill_green, al=al_r, border=b_all)
put(ws, "E4", "Δ (economia B)", font=f_bold, fill=fill_zebra, al=al_r, border=b_all)
concl = [
 ("Custo total do investimento após a incorporação", f"={INC}!C12*2", f"={INC}!C16*2", "=D5-C5"),
 ("IR no cash out (etapa 3)", f"={CO}!F24", f"={CO}!G24", "=C6-D6"),
 ("Custo remanescente para o passo F", f"={CO}!C27*2", f"={CO}!D27*2", "=D7-C7"),
 ("IR potencial na alienação futura (teto da economia: 34% × Δ custo)", f"=MAX(0,C7*{P}!C8)", "", "=C7*0"),
]
r = 5
put(ws, f"B{r}", concl[0][0], border=b_all)
put(ws, f"C{r}", concl[0][1], nf=NF, border=b_all)
put(ws, f"D{r}", concl[0][2], nf=NF, border=b_all)
put(ws, f"E{r}", concl[0][3], nf=NF, border=b_all)
r += 1
put(ws, f"B{r}", concl[1][0], font=f_bold, border=b_all)
put(ws, f"C{r}", concl[1][1], font=f_bold, nf=NF, border=b_all)
put(ws, f"D{r}", concl[1][2], font=f_bold, nf=NF, border=b_all)
put(ws, f"E{r}", concl[1][3], font=f_bold, nf=NF, border=b_all)
r += 1
put(ws, f"B{r}", concl[2][0], border=b_all)
put(ws, f"C{r}", concl[2][1], nf=NF, border=b_all)
put(ws, f"D{r}", concl[2][2], nf=NF, border=b_all)
put(ws, f"E{r}", concl[2][3], nf=NF, border=b_all)
r += 1
put(ws, f"B{r}", "Economia futura máxima da Hip. B (34% × Δ custo remanescente)", font=f_bold, border=b_all)
put(ws, f"C{r}", "", border=b_all)
put(ws, f"D{r}", "", border=b_all)
put(ws, f"E{r}", f"=E7*{P}!C8", font=f_bold, nf=NF, border=b_all)
r += 1
put(ws, f"B{r}", "BENEFÍCIO TOTAL DA TESE DO CUSTO \"GORDINHO\"", font=f_bold, fill=fill_green, border=b_all)
put(ws, f"C{r}", "", border=b_all); ws[f"C{r}"].fill = fill_green
put(ws, f"D{r}", "", border=b_all); ws[f"D{r}"].fill = fill_green
put(ws, f"E{r}", "=E6+E8", font=f_bold, nf=NF, fill=fill_green, border=b_all)
r += 1
put(ws, f"B{r}", "Check: 34% × parcela incorporada total (689,94)", font=f_it_gray, border=b_all)
put(ws, f"C{r}", "", border=b_all); put(ws, f"D{r}", "", border=b_all)
put(ws, f"E{r}", f"=E9-{INC}!D15*{P}!C8", font=f_it_gray, nf=NF, border=b_all)

bar(ws, f"B{r+2}:E{r+2}", "LEITURA E RISCOS")
leitura = [
 ("Onde nasce a diferença", "A parcela incorporada (344,97 por holding; 689,94 no total) = reflexo do PL contábil da CNP (419,37) − diluição de 13,8% (74,40). Compõe custo na Hip. B; não compõe na Hip. A."),
 ("Valor da tese", "34% × 689,94 = 234,58 no total — sendo 71,30 já no cash out (fração alienada de 30,39%) e até 163,28 na alienação futura (fração retida de 69,61%)."),
 ("Riscos por camada", "Perda de diluição (33,05): menor controvérsia. Reflexo do PL contábil da CNP (472,53): neutralidade defensável (MEP/variação de participação). Reflexo da parcela AVJ implícita (250,46): maior risco de requalificação (AVJ Reflexo — tributação na alienação)."),
 ("Jurisprudência", "WTorre (1402-004.537/2020) suporta a exclusão definitiva (\"dá custo\"); CPFL, Litela e Litel (2024) apontam diferimento — tributa quando o ganho migra de ORA/reserva para a DRE. A conta de PL da contrapartida (ORA x reserva) segue indefinida."),
 ("Quem decide o enquadramento", "O auditor da Embracon (Deloitte) define CPC 15 x CPC 19. Sob CPC 19, o AVJ fica carimbado a 34% em subconta e a Hip. B deixa de existir; \"cai a tese do MEP\"."),
]
r += 3
for t, d in leitura:
    put(ws, f"B{r}", t, font=f_bold, al=al_lt, border=b_all)
    ws.merge_cells(f"C{r}:E{r}")
    put(ws, f"C{r}", d, font=f_it, al=al_lt, border=b_all)
    for cc in "DE": ws[f"{cc}{r}"].border = b_all
    if r % 2 == 0:
        for cc in "BCDE": ws[f"{cc}{r}"].fill = fill_zebra
    ws.row_dimensions[r].height = 30
    r += 1

wb.save(OUT)
print("saved", OUT)
